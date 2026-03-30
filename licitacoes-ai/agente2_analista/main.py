"""Agente 2 — Analista: baixa PDF/XLSX, analisa edital+TR+anexos, gera parecer Go/No-Go."""
import json
import logging
import time
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.database import (
    get_edital, atualizar_status_edital, registrar_execucao
)
from shared.utils import derivar_esfera
from agente1_monitor.pncp_client import buscar_arquivos_compra
from config.settings import EDITAIS_DIR
from agente2_analista.pdf_extractor import download_pdf, extract_smart, MAX_CHARS
from agente2_analista.table_extractor import extrair_postos_tabela
from agente2_analista.edital_parser import extrair_dados_estruturados
from agente2_analista.viability_checker import verificar_viabilidade

log = logging.getLogger("agente2_analista")


# ─── Fase 1: Baixar todos os documentos ───────────────────────────

def _baixar_todos_arquivos(pncp_id: str, cnpj: str, ano: str, seq: str) -> dict:
    """Baixa edital, TR e todos os anexos do PNCP.

    Returns:
        {
            "edital_pdf": Path | None,
            "tr_pdf": Path | None,
            "planilhas_xlsx": [Path, ...],
            "outros_pdfs": [Path, ...],
            "total_arquivos": int,
        }
    """
    result = {
        "edital_pdf": None,
        "tr_pdf": None,
        "planilhas_xlsx": [],
        "outros_pdfs": [],
        "total_arquivos": 0,
    }

    try:
        arquivos = buscar_arquivos_compra(cnpj, int(ano), int(seq))
    except Exception as e:
        log.warning(f"Erro ao buscar arquivos PNCP: {e}")
        return result

    if not arquivos:
        log.warning(f"Nenhum arquivo encontrado para {pncp_id}")
        return result

    result["total_arquivos"] = len(arquivos)
    log.info(f"Encontrados {len(arquivos)} arquivos no PNCP")

    EDITAIS_DIR.mkdir(parents=True, exist_ok=True)

    for i, arq in enumerate(arquivos):
        url = arq.get("url") or arq.get("uri") or ""
        titulo = (arq.get("titulo", "") or "").lower()
        if not url:
            continue

        try:
            # Classificar tipo do arquivo
            is_edital = "edital" in titulo and "tr" not in titulo
            is_tr = any(k in titulo for k in ("termo_de_referencia", "termo de referencia", "termo_referencia", "tr_", "_tr"))
            is_planilha = any(k in titulo for k in ("planilha", ".xlsx", ".xls", "custo", "orcamento", "preco"))

            if is_edital and not result["edital_pdf"]:
                pdf_path = download_pdf(url, f"{pncp_id}.pdf")
                result["edital_pdf"] = pdf_path
                log.info(f"  Edital: {titulo}")

            elif is_tr and not result["tr_pdf"]:
                pdf_path = download_pdf(url, f"{pncp_id}_TR.pdf")
                result["tr_pdf"] = pdf_path
                log.info(f"  TR: {titulo}")

            elif is_planilha:
                # Planilhas XLSX — baixar separado
                ext = ".xlsx" if ".xlsx" in titulo or ".xls" in titulo else ".pdf"
                fname = f"{pncp_id}_anexo_{i}{ext}"
                path = _download_file(url, fname)
                if path:
                    result["planilhas_xlsx"].append(path)
                    log.info(f"  Planilha: {titulo}")

            else:
                # Outros documentos
                fname = f"{pncp_id}_doc_{i}.pdf"
                path = download_pdf(url, fname)
                result["outros_pdfs"].append(path)
                log.info(f"  Outro: {titulo}")

        except Exception as e:
            log.warning(f"  Erro baixando '{titulo}': {e}")

    # Se não identificou edital específico, usa o primeiro PDF
    if not result["edital_pdf"] and result["outros_pdfs"]:
        result["edital_pdf"] = result["outros_pdfs"].pop(0)
        log.info("Usando primeiro PDF como edital principal")

    # Fallback: busca arquivos locais ja baixados (ex: do site do orgao)
    import glob as _glob
    local_pattern = str(EDITAIS_DIR / f"{pncp_id}_*")
    for fpath in sorted(_glob.glob(local_pattern)):
        fp = Path(fpath)
        fname_lower = fp.name.lower()

        if fp.suffix.lower() == ".pdf":
            is_tr_local = any(k in fname_lower for k in (
                "termo_de_referencia", "termo de referencia", "termo-de-referencia",
                "_tr_", "_tr-", "_tr.", "-tr.", "-tr-", "anexo-i-tr", "anexo_i_tr",
                "anexo-i-", "anexoi-tr",
            ))
            is_edital_local = "edital" in fname_lower and not is_tr_local
            is_anexo = "anexo" in fname_lower

            if is_tr_local and not result["tr_pdf"]:
                result["tr_pdf"] = fp
                log.info(f"  TR (local): {fp.name}")
            elif is_edital_local and not result["edital_pdf"]:
                result["edital_pdf"] = fp
                log.info(f"  Edital (local): {fp.name}")
            elif is_anexo:
                if fp not in result["outros_pdfs"]:
                    result["outros_pdfs"].append(fp)
                    log.info(f"  Anexo (local): {fp.name}")

        elif fp.suffix.lower() in (".xlsx", ".xls"):
            if fp not in result["planilhas_xlsx"]:
                result["planilhas_xlsx"].append(fp)
                log.info(f"  Planilha (local): {fp.name}")

    if not result["edital_pdf"]:
        # Ultimo fallback: arquivo principal sem sufixo
        main_pdf = EDITAIS_DIR / f"{pncp_id}.pdf"
        if main_pdf.exists():
            result["edital_pdf"] = main_pdf
            log.info(f"  Edital (fallback): {main_pdf.name}")

    return result


def _download_file(url: str, filename: str) -> Path | None:
    """Baixa qualquer arquivo (XLSX, PDF, etc)."""
    import httpx
    save_path = EDITAIS_DIR / filename
    if save_path.exists():
        return save_path
    try:
        with httpx.Client(timeout=60, follow_redirects=True) as client:
            resp = client.get(url)
            resp.raise_for_status()
            save_path.write_bytes(resp.content)
            log.info(f"Arquivo salvo: {save_path.name} ({len(resp.content)/1024:.0f} KB)")
            return save_path
    except Exception as e:
        log.warning(f"Erro baixando {url}: {e}")
        return None


# ─── Fase 2: Analisar edital (requisitos, habilitação, prazos) ────

def _analisar_edital_pdf(pdf_path: Path, pncp_id: str, tr_path: Path = None) -> tuple[str, dict]:
    """Extrai texto do edital + TR e analisa via LLM numa única chamada.

    Returns:
        (texto_edital, analise_dados)
    """
    result = extract_smart(pdf_path)
    texto = result["text"]
    log.info(f"Edital: {result['pages']} pgs, {len(texto)} chars")

    # Extrair TR se disponível
    texto_tr = None
    if tr_path and tr_path.exists():
        tr_result = extract_smart(tr_path)
        texto_tr = tr_result["text"]
        log.info(f"TR: {tr_result['pages']} pgs, {len(texto_tr)} chars")

    # Análise LLM única — edital + TR juntos
    analise = extrair_dados_estruturados(texto, pncp_id, texto_tr=texto_tr)
    return texto, analise


# ─── Fase 3: Analisar TR (postos, CCT, benefícios) ───────────────

def _analisar_tr_pdf(tr_path: Path, analise_edital: dict, pncp_id: str) -> dict:
    """Analisa o Termo de Referência com foco em postos e CCT.

    O TR geralmente tem os detalhes que o edital não tem:
    - Postos de trabalho com quantidades exatas
    - CCT aplicável com pisos salariais
    - Benefícios obrigatórios (VT, VA, uniforme, etc)
    - Jornadas de trabalho detalhadas
    - Modelo de planilha de custos
    """
    from shared.llm_client import ask_claude_json

    result = extract_smart(tr_path)
    texto_tr = result["text"]
    log.info(f"TR: {result['pages']} pgs, {len(texto_tr)} chars")

    # Extração de tabelas primeiro (gratuito)
    postos_tabela = extrair_postos_tabela(tr_path)
    if postos_tabela:
        log.info(f"TR tabelas: {len(postos_tabela)} postos extraídos gratuitamente")

    # Análise LLM específica do TR
    prompt_tr = f"""Analise este TERMO DE REFERÊNCIA de licitação.

FOCO PRINCIPAL: extraia com PRECISÃO MÁXIMA os dados para precificação.

Responda APENAS com JSON válido:
{{
  "postos_trabalho": [
    {{
      "funcao": "<nome exato do cargo/posto>",
      "quantidade": <int>,
      "jornada": "<44h|40h|36h|30h|12x36|escala>",
      "salario_base": <float se mencionado, null se não>,
      "cbo": "<código CBO se mencionado>",
      "escolaridade": "<fundamental|medio|tecnico|superior>",
      "adicional_periculosidade": <true|false>,
      "adicional_insalubridade": <true|false>,
      "adicional_noturno": <true|false>,
      "descricao": "<atividades principais>"
    }}
  ],
  "cct_aplicavel": {{
    "sindicato_patronal": "<ex: SEAC-RJ>",
    "sindicato_laboral": "<ex: SIEMACO-RIO>",
    "numero_registro": "<ex: RJ001061/2025>",
    "vigencia": "<ex: 01/03/2025 a 28/02/2026>",
    "categoria": "<limpeza|vigilancia|administrativo|facilities|outros>"
  }},
  "beneficios_obrigatorios": {{
    "vale_transporte": "<descricao ou valor>",
    "vale_alimentacao": "<descricao ou valor>",
    "cesta_basica": "<descricao ou valor>",
    "plano_saude": "<descricao ou valor>",
    "seguro_vida": "<descricao ou valor>",
    "uniforme": "<descricao ou valor>",
    "outros": ["<outros benefícios>"]
  }},
  "modelo_planilha": "<IN05/2017|IN18/2022|SEGES|proprio_orgao|null>",
  "prazo_contrato_meses": <int>,
  "local_execucao": "<endereço/cidade>",
  "horario_funcionamento": "<ex: 07h às 19h>",
  "encargos_especificos": {{
    "rat_sat": <float se mencionado>,
    "desonerado": <true|false|null>,
    "insalubridade_grau": "<minimo|medio|maximo|null>",
    "periculosidade": <true|false>
  }},
  "observacoes_precificacao": [
    "<qualquer info relevante para montar a planilha de custos>"
  ]
}}

TERMO DE REFERÊNCIA:
{texto_tr[:25000]}"""

    try:
        tr_dados = ask_claude_json(
            system="Você é um especialista em precificação de licitações. Extraia dados para montar planilha de custos.",
            user=prompt_tr,
            max_tokens=4096,
            agente="analista_tr",
            pncp_id=pncp_id,
        )
    except Exception as e:
        log.warning(f"Erro na análise LLM do TR: {e}")
        tr_dados = {}

    # Postos da tabela têm prioridade (mais precisos)
    if postos_tabela and len(postos_tabela) > 0:
        tr_dados["postos_trabalho"] = postos_tabela
        tr_dados["_postos_fonte"] = "tabela_pdf"
    elif tr_dados.get("postos_trabalho"):
        tr_dados["_postos_fonte"] = "llm_tr"

    return tr_dados


# ─── Fase 4: Analisar planilhas XLSX do órgão ────────────────────

def _analisar_planilhas_anexas(planilhas: list[Path], pncp_id: str) -> dict:
    """Lê planilhas XLSX anexas ao edital para entender o modelo do órgão.

    Muitos órgãos fornecem uma planilha modelo que a empresa deve preencher.
    Essa planilha define a estrutura (módulos, encargos, benefícios).
    """
    if not planilhas:
        return {}

    resultado = {
        "modelo_encontrado": False,
        "abas": [],
        "estrutura_modulos": [],
        "observacoes": [],
    }

    for xlsx_path in planilhas:
        if not xlsx_path.suffix.lower() in ('.xlsx', '.xls'):
            continue

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
            resultado["modelo_encontrado"] = True
            resultado["arquivo"] = xlsx_path.name

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                aba_info = {
                    "nome": sheet_name,
                    "linhas": ws.max_row,
                    "colunas": ws.max_column,
                    "preview": [],
                }
                # Pegar primeiras 20 linhas para entender a estrutura
                for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=False):
                    row_data = []
                    for c in row:
                        if c.value is not None:
                            row_data.append(f"{c.coordinate}={c.value}")
                    if row_data:
                        aba_info["preview"].append(" | ".join(row_data[:6]))

                resultado["abas"].append(aba_info)

            log.info(f"Planilha modelo: {xlsx_path.name}, {len(wb.sheetnames)} abas: {wb.sheetnames}")
            break  # Usa a primeira planilha encontrada

        except Exception as e:
            log.warning(f"Erro lendo planilha {xlsx_path.name}: {e}")

    return resultado


# ─── Pipeline principal ──────────────────────────────────────────

def analisar_edital(pncp_id: str) -> dict:
    """Pipeline completo de análise em 4 fases.

    Fase 1: Baixar todos os documentos (edital, TR, planilhas, anexos)
    Fase 2: Analisar edital (requisitos, habilitação, prazos)
    Fase 3: Analisar TR (postos, CCT, benefícios, encargos)
    Fase 4: Analisar planilha do órgão (modelo de custos)

    Returns:
        {"parecer": str, "analise": dict, "viabilidade": dict}
    """
    start = time.time()

    # 0. Busca edital no banco
    edital = get_edital(pncp_id)
    if not edital:
        raise ValueError(f"Edital {pncp_id} não encontrado no banco")

    atualizar_status_edital(pncp_id, "analisando")
    log.info(f"{'='*60}")
    log.info(f"ANALISANDO: {pncp_id}")
    log.info(f"Objeto: {edital['objeto'][:100]}")
    log.info(f"{'='*60}")

    # Parse do pncp_id
    parts = pncp_id.split("-")
    is_manual = pncp_id.startswith("MANUAL")

    if not is_manual and len(parts) >= 3:
        cnpj, ano, seq = parts[0], parts[1], parts[2]
    elif is_manual:
        cnpj, ano, seq = "", "", ""
    else:
        raise ValueError(f"pncp_id inválido: {pncp_id}")

    # ─── FASE 1: Baixar todos os documentos ───────────────────
    log.info("FASE 1: Baixando documentos...")

    if not is_manual:
        docs = _baixar_todos_arquivos(pncp_id, cnpj, ano, seq)
    else:
        docs = {"edital_pdf": None, "tr_pdf": None, "planilhas_xlsx": [], "outros_pdfs": [], "total_arquivos": 0}

    log.info(f"  Edital: {'OK' if docs['edital_pdf'] else 'N/A'}")
    log.info(f"  TR: {'OK' if docs['tr_pdf'] else 'N/A'}")
    log.info(f"  Planilhas: {len(docs['planilhas_xlsx'])}")
    log.info(f"  Outros: {len(docs['outros_pdfs'])}")

    # ─── FASE 2: Analisar edital ──────────────────────────────
    log.info("FASE 2: Analisando edital...")

    texto_edital = ""
    analise_dados = {}

    if docs["edital_pdf"]:
        texto_edital, analise_dados = _analisar_edital_pdf(docs["edital_pdf"], pncp_id, tr_path=docs.get("tr_pdf"))
    else:
        # Sem PDF — usa dados básicos da API
        texto_edital = f"""EDITAL (dados da API PNCP — sem PDF disponível):
Órgão: {edital.get('orgao_nome', 'N/I')}
Objeto: {edital.get('objeto', 'N/I')}
Valor estimado: R$ {edital.get('valor_estimado', 'N/I')}
UF: {edital.get('uf', 'N/I')}
Modalidade: {edital.get('modalidade', 'N/I')}
Data abertura: {edital.get('data_abertura', 'N/I')}"""

        analise_dados = {
            "objeto_detalhado": edital.get("objeto", ""),
            "postos_trabalho": [],
            "habilitacao": {},
            "riscos_identificados": [],
            "oportunidades": [],
        }

    log.info(f"  Postos edital: {len(analise_dados.get('postos_trabalho', []))}")
    log.info(f"  CCT: {analise_dados.get('cct_aplicavel', {}).get('sindicato', 'N/I')}")

    # ─── FASE 3: Analisar Termo de Referência ─────────────────
    log.info("FASE 3: Analisando Termo de Referência...")

    tr_dados = {}
    if docs["tr_pdf"]:
        tr_dados = _analisar_tr_pdf(docs["tr_pdf"], analise_dados, pncp_id)

        # TR complementa/substitui dados do edital
        if tr_dados.get("postos_trabalho"):
            analise_dados["postos_trabalho"] = tr_dados["postos_trabalho"]
            analise_dados["_postos_fonte"] = tr_dados.get("_postos_fonte", "tr")
            log.info(f"  Postos do TR: {len(tr_dados['postos_trabalho'])}")

        if tr_dados.get("cct_aplicavel"):
            analise_dados["cct_aplicavel"] = tr_dados["cct_aplicavel"]
            log.info(f"  CCT do TR: {tr_dados['cct_aplicavel']}")

        if tr_dados.get("beneficios_obrigatorios"):
            analise_dados["beneficios_obrigatorios"] = tr_dados["beneficios_obrigatorios"]

        if tr_dados.get("modelo_planilha"):
            analise_dados["modelo_planilha"] = tr_dados["modelo_planilha"]

        if tr_dados.get("encargos_especificos"):
            analise_dados["encargos_especificos"] = tr_dados["encargos_especificos"]

        if tr_dados.get("prazo_contrato_meses"):
            analise_dados["prazo_contrato_meses"] = tr_dados["prazo_contrato_meses"]

        if tr_dados.get("observacoes_precificacao"):
            analise_dados["observacoes_precificacao"] = tr_dados["observacoes_precificacao"]
    else:
        log.info("  TR não disponível — usando dados do edital")

    # Tenta extrair postos de tabelas — prioridade: TR > outros PDFs > edital
    # NUNCA usar o edital principal primeiro (pode ter itens de compra/materiais)
    if not analise_dados.get("postos_trabalho") or len(analise_dados["postos_trabalho"]) == 0:
        pdfs_busca = []
        if docs.get("tr_pdf"):
            pdfs_busca.append(docs["tr_pdf"])
        pdfs_busca.extend(docs.get("outros_pdfs", []))
        # Edital principal só como último recurso
        if docs.get("edital_pdf"):
            pdfs_busca.append(docs["edital_pdf"])

        for pdf_path in pdfs_busca:
            if pdf_path and pdf_path.exists():
                postos = extrair_postos_tabela(pdf_path)
                if postos:
                    # Validar: postos reais têm nomes curtos de profissão, não descrições de materiais
                    nomes_validos = [p for p in postos if len(p.get("funcao_display", "")) < 80]
                    if nomes_validos and len(nomes_validos) >= len(postos) * 0.5:
                        analise_dados["postos_trabalho"] = postos
                        analise_dados["_postos_fonte"] = f"tabela_{pdf_path.name}"
                        log.info(f"  Postos de tabela ({pdf_path.name}): {len(postos)}")
                        break
                    else:
                        log.warning(f"  Descartando {len(postos)} 'postos' de {pdf_path.name} — parecem itens de compra")

    # ─── FASE 4: Analisar planilhas XLSX do órgão ─────────────
    log.info("FASE 4: Analisando planilhas do órgão...")

    if docs["planilhas_xlsx"]:
        modelo = _analisar_planilhas_anexas(docs["planilhas_xlsx"], pncp_id)
        if modelo.get("modelo_encontrado"):
            analise_dados["planilha_modelo_orgao"] = modelo
            log.info(f"  Modelo do órgão: {modelo.get('arquivo', 'N/I')}, {len(modelo.get('abas', []))} abas")
        else:
            log.info("  Nenhum modelo XLSX encontrado")
    else:
        log.info("  Sem planilhas anexas")

    # ─── FASE 5: Ranking de empresas + viabilidade ────────────
    log.info("FASE 5: Avaliando viabilidade e ranking...")

    from agente2_analista.viability_checker import rankear_empresas
    esfera = derivar_esfera(cnpj) if not is_manual else "estadual"
    atestados_exigidos = analise_dados.get("habilitacao", {}).get("qualificacao_tecnica", [])

    ranking = rankear_empresas(
        objeto=edital.get("objeto", ""),
        uf_edital=edital.get("uf", ""),
        esfera=esfera,
        data_abertura=edital.get("data_abertura"),
        data_encerramento=edital.get("data_encerramento"),
        atestados_exigidos=atestados_exigidos,
    )
    analise_dados["empresas_ranking"] = ranking

    melhor = next((r for r in ranking if r["viavel"]), None)
    empresa = melhor["nome"].lower() if melhor else edital.get("empresa_sugerida", "manutec")

    viabilidade = verificar_viabilidade(
        empresa_nome=empresa,
        uf_edital=edital.get("uf", ""),
        esfera=esfera,
        data_abertura=edital.get("data_abertura"),
        data_encerramento=edital.get("data_encerramento"),
        atestados_exigidos=atestados_exigidos,
    )
    parecer = viabilidade["parecer"]

    # ─── Salvar no banco ──────────────────────────────────────
    analise_json = json.dumps(analise_dados, ensure_ascii=False, default=str)
    requisitos_json = json.dumps(
        analise_dados.get("habilitacao", {}), ensure_ascii=False
    )

    atualizar_status_edital(
        pncp_id,
        status=parecer,
        analise_json=analise_json,
        parecer=parecer,
        motivo_nogo=viabilidade.get("motivo_nogo"),
        requisitos_habilitacao=requisitos_json,
        empresa_sugerida=empresa,
    )

    duracao = time.time() - start
    registrar_execucao(
        agente="analista",
        pncp_id=pncp_id,
        status="sucesso",
        duracao_seg=duracao,
    )

    # ─── Resumo ───────────────────────────────────────────────
    n_postos = len(analise_dados.get("postos_trabalho", []))
    total_func = sum(p.get("quantidade", 1) for p in analise_dados.get("postos_trabalho", []))
    cct = analise_dados.get("cct_aplicavel", {})
    cct_info = cct.get("sindicato_patronal") or cct.get("sindicato") or "N/I"

    log.info(f"{'='*60}")
    log.info(f"RESULTADO: {parecer.upper()}")
    log.info(f"  Empresa: {empresa}")
    log.info(f"  Postos: {n_postos} cargos, {total_func} funcionários")
    log.info(f"  CCT: {cct_info}")
    log.info(f"  Fonte postos: {analise_dados.get('_postos_fonte', 'N/I')}")
    log.info(f"  Docs baixados: {docs['total_arquivos']}")
    log.info(f"  Duração: {duracao:.1f}s")
    log.info(f"{'='*60}")

    return {
        "parecer": parecer,
        "analise": analise_dados,
        "viabilidade": viabilidade,
        "docs": {
            "edital": str(docs["edital_pdf"]) if docs["edital_pdf"] else None,
            "tr": str(docs["tr_pdf"]) if docs["tr_pdf"] else None,
            "planilhas": [str(p) for p in docs["planilhas_xlsx"]],
            "total": docs["total_arquivos"],
        },
    }


if __name__ == "__main__":
    import sys as _sys
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    )

    if len(_sys.argv) < 2:
        print("Uso: python -m agente2_analista.main <pncp_id>")
        print("Ex:  python -m agente2_analista.main 42498733000148-2026-442")
        _sys.exit(1)

    result = analisar_edital(_sys.argv[1])
    print(f"\nParecer: {result['parecer']}")
    print(f"Docs: {result['docs']}")
    print(json.dumps(result["analise"], indent=2, ensure_ascii=False)[:3000])
