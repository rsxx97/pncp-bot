"""Preenche planilha-template do orgao com dados calculados."""
import copy
import logging
from pathlib import Path

import openpyxl

log = logging.getLogger("template_filler")


def detectar_template(editais_dir: Path, pncp_id: str) -> Path | None:
    """Detecta se existe planilha-template do orgao nos arquivos baixados."""
    import glob
    pattern = str(editais_dir / f"{pncp_id}*")
    for fpath in sorted(glob.glob(pattern)):
        fp = Path(fpath)
        if fp.suffix.lower() in (".xlsx", ".xls"):
            fname = fp.name.lower()
            keywords = ["planilha", "mao_de_obra", "mao de obra", "modelo", "custos", "formacao"]
            if any(k in fname for k in keywords):
                log.info(f"Template encontrado: {fp.name}")
                return fp
    return None


def preencher_template(
    template_path: Path,
    postos: list[dict],
    empresa_info: dict,
    output_path: Path,
) -> Path:
    """Preenche template do orgao com dados dos postos.

    Args:
        template_path: Caminho do template XLSX do orgao
        postos: Lista de dicts com dados calculados por cargo:
            [{funcao, salario, quantidade, jornada, vt, va, uniformes,
              insalubridade, periculosidade, noturno, ...}]
        empresa_info: Dict com dados da empresa:
            {nome, regime, ci_pct, lucro_pct, pis_pct, cofins_pct, iss_pct}
        output_path: Caminho de saida

    Returns:
        Path do arquivo gerado
    """
    wb = openpyxl.load_workbook(str(template_path))

    # Template tem 1 aba "Cargo" - duplica para cada posto
    template_sheet_name = wb.sheetnames[0]
    template_ws = wb[template_sheet_name]

    # Mapeia celulas do template
    cell_map = _mapear_celulas(template_ws)

    if not cell_map:
        log.warning("Nao consegui mapear celulas do template, usando builder padrao")
        return None

    # Remove a aba original
    sheets_created = []

    for i, posto in enumerate(postos):
        funcao = posto.get("funcao", f"Cargo {i+1}").upper()
        # Copia a aba template
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = funcao[:31]  # Excel limita 31 chars

        # Preenche dados
        _preencher_aba(new_ws, cell_map, posto, empresa_info)
        sheets_created.append(new_ws)

    # Remove template original
    wb.remove(template_ws)

    # Adiciona aba RESUMO
    _criar_resumo(wb, postos, empresa_info)

    # Adiciona aba BREAK-EVEN
    _criar_breakeven(wb, postos, empresa_info)

    wb.save(str(output_path))
    log.info(f"Planilha gerada: {output_path} ({len(postos)} cargos)")
    return output_path


def _mapear_celulas(ws) -> dict:
    """Mapeia posicoes das celulas-chave no template."""
    cell_map = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value or "").lower().strip()

            # Modulo 1
            if "sal" in val and "base" in val:
                cell_map["salario_base"] = (cell.row, "D")
            if "periculosidade" in val:
                cell_map["periculosidade"] = (cell.row, "D")
            if "insalubridade" in val:
                cell_map["insalubridade"] = (cell.row, "D")
            if "noturno" in val and "hora" not in val:
                cell_map["noturno"] = (cell.row, "D")

            # Modulo 2.3 - Beneficios
            if "transporte" in val:
                cell_map["vt"] = (cell.row, "D")
            if ("alimenta" in val or "refei" in val) and "aux" in val:
                cell_map["va"] = (cell.row, "D")
            if "assist" in val and "med" in val:
                cell_map["assist_medica"] = (cell.row, "D")
            if "seguro" in val and "vida" in val:
                cell_map["seguro_vida"] = (cell.row, "D")

            # SAT/RAT
            if "sat" in val or "rat" in val:
                cell_map["sat_rat"] = (cell.row, "C")

            # Modulo 5
            if "uniformes" in val or "uniforme" in val:
                cell_map["uniformes"] = (cell.row, "D")
            if "materiais" in val or "material" in val:
                cell_map["materiais"] = (cell.row, "D")
            if "equipamentos" in val or "equipamento" in val:
                cell_map["equipamentos"] = (cell.row, "D")

            # Modulo 6
            if "custos indiretos" in val and "tribut" not in val and "lucro" not in val:
                cell_map["ci"] = (cell.row, "C")
                cell_map["ci_val"] = (cell.row, "D")
            if "lucro" == val.strip() or (val.startswith("lucro") and len(val) < 10):
                cell_map["lucro"] = (cell.row, "C")
                cell_map["lucro_val"] = (cell.row, "D")
            if "iss" in val and "c." in val:
                cell_map["iss"] = (cell.row, "C")
            if "cofins" in val and "c." in val:
                cell_map["cofins"] = (cell.row, "C")
            if "pis" in val and "c." in val:
                cell_map["pis"] = (cell.row, "C")

            # Header info
            if "cbo" in val:
                cell_map["cbo"] = (cell.row, "D" if ":" in val else "C")
            if "categoria" in val and "profissional" in val:
                cell_map["categoria"] = (cell.row, "D" if ":" in val else "C")
            if "quantidade" in val and "empregad" in val:
                cell_map["quantidade"] = (cell.row, "C")

    return cell_map


def _preencher_aba(ws, cell_map: dict, posto: dict, empresa: dict):
    """Preenche uma aba do template com dados do posto.

    O template do orgao tem formulas encadeadas. Precisamos preencher
    apenas as celulas de INPUT — as formulas calculam o resto.
    """
    # Abordagem: varre todas as celulas e identifica os campos de input
    # (celulas com valor 0 ou None que sao referenciadas por formulas)

    salario = posto.get("salario", 0)

    # Preenche por busca inteligente de celulas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value or "").lower().strip()

            # Salario Normativo (D7 tipicamente) — o D13 faz =D7
            if ("sal" in val and "normativo" in val) or ("sal" in val and "categoria" in val):
                # O valor fica na mesma linha, coluna D
                target = ws[f"D{cell.row}"]
                if target.value is None or target.value == 0:
                    target.value = salario

            # Categoria profissional
            if "categoria profissional" in val and "vinculada" in val:
                target = ws[f"D{cell.row}"]
                if target.value is None or target.value == 0 or target.value == "":
                    target.value = posto.get("funcao", "").upper()

    # Preenche celulas mapeadas por keyword
    def set_cell(key, value):
        if key in cell_map:
            r, col = cell_map[key]
            c = ws[f"{col}{r}"]
            # So preenche se a celula nao tem formula
            if c.value is None or c.value == 0 or (isinstance(c.value, str) and not c.value.startswith("=")):
                c.value = value

    # Periculosidade/Insalubridade como % na coluna C
    if posto.get("periculosidade", 0) > 0:
        set_cell("periculosidade", posto["periculosidade"])
    if posto.get("insalubridade", 0) > 0:
        set_cell("insalubridade", posto["insalubridade"])
    if posto.get("adicional_noturno", 0) > 0:
        set_cell("noturno", posto["adicional_noturno"])

    # SAT/RAT %
    set_cell("sat_rat", empresa.get("sat_rat_pct", 0.02))

    # Beneficios (valores absolutos na coluna D)
    set_cell("vt", posto.get("vt", 0))
    set_cell("va", posto.get("va", 0))
    set_cell("assist_medica", posto.get("assist_medica", 0))
    set_cell("seguro_vida", posto.get("seguro_vida", 0))

    # Insumos
    set_cell("uniformes", posto.get("uniformes", 0))
    set_cell("materiais", posto.get("materiais", 0))
    set_cell("equipamentos", posto.get("equipamentos", 0))

    # Modulo 6 — percentuais na coluna C
    set_cell("ci", empresa.get("ci_pct", 0.03))
    set_cell("lucro", empresa.get("lucro_pct", 0.03))

    # ISS/COFINS/PIS — busca direta se mapeador nao pegou
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value or "").lower().strip()
            if "iss" in val and cell.column_letter == "B":
                c = ws[f"C{cell.row}"]
                if c.value == 0 or c.value is None:
                    c.value = empresa.get("iss_pct", 0.02)
            if "cofins" in val and cell.column_letter == "B":
                c = ws[f"C{cell.row}"]
                if c.value == 0 or c.value is None:
                    c.value = empresa.get("cofins_pct", 0.0415)
            if "pis" in val and cell.column_letter == "B":
                c = ws[f"C{cell.row}"]
                if c.value == 0 or c.value is None:
                    c.value = empresa.get("pis_pct", 0.0005)
            # SAT/RAT
            if ("sat" in val or "rat" in val) and cell.column_letter in ("A", "B"):
                c = ws[f"C{cell.row}"]
                if c.value == 0 or c.value is None:
                    c.value = empresa.get("sat_rat_pct", 0.02)

    # Quantidade de empregados
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value or "").lower().strip()
            if "quantidade" in val and "empregad" in val:
                target = ws[f"C{cell.row}"]
                if target.value is not None:
                    target.value = posto.get("quantidade", 1)
                else:
                    target = ws[f"D{cell.row}"]
                    target.value = posto.get("quantidade", 1)
    set_cell("iss", empresa.get("iss_pct", 0.02))
    set_cell("cofins", empresa.get("cofins_pct", 0.0415))
    set_cell("pis", empresa.get("pis_pct", 0.0005))

    # Quantidade
    set_cell("quantidade", posto.get("quantidade", 1))

    # CBO
    if posto.get("cbo"):
        set_cell("cbo", posto["cbo"])
    if posto.get("funcao"):
        set_cell("categoria", posto["funcao"].upper())


def _criar_resumo(wb, postos, empresa_info):
    """Cria aba RESUMO consolidando todos os cargos."""
    ws = wb.create_sheet("RESUMO")

    orgao = empresa_info.get("orgao", "")
    pregao = empresa_info.get("pregao", "")
    nome_empresa = empresa_info.get("nome", "")

    ws["A1"] = f"{orgao} - {pregao} ({nome_empresa})"
    ws["A2"] = f"Regime: {empresa_info.get('regime', 'Lucro Real')}"

    ws["A4"] = "Item"
    ws["B4"] = "Cargo"
    ws["C4"] = "Postos"
    ws["D4"] = "R$/Emp/Mes"
    ws["E4"] = "R$ Mensal"
    ws["F4"] = "R$ Anual"

    total_postos = 0
    total_mensal = 0

    for i, p in enumerate(postos):
        row = 5 + i
        custo_emp = p.get("custo_total_mensal", 0)
        qtd = p.get("quantidade", 1)
        mensal = custo_emp * qtd
        anual = mensal * 12

        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = p.get("funcao", "").upper()
        ws[f"C{row}"] = qtd
        ws[f"D{row}"] = round(custo_emp, 2)
        ws[f"E{row}"] = round(mensal, 2)
        ws[f"F{row}"] = round(anual, 2)

        total_postos += qtd
        total_mensal += mensal

    row_total = 5 + len(postos)
    ws[f"B{row_total}"] = "TOTAL"
    ws[f"C{row_total}"] = total_postos
    ws[f"E{row_total}"] = round(total_mensal, 2)
    ws[f"F{row_total}"] = round(total_mensal * 12, 2)

    teto = empresa_info.get("valor_edital", 0)
    proposta = total_mensal * 12

    ws[f"B{row_total + 2}"] = "TETO EDITAL (anual):"
    ws[f"F{row_total + 2}"] = teto
    ws[f"B{row_total + 3}"] = "NOSSA PROPOSTA:"
    ws[f"F{row_total + 3}"] = round(proposta, 2)
    if teto > 0:
        ws[f"B{row_total + 4}"] = "DESCONTO:"
        ws[f"F{row_total + 4}"] = f"{((teto - proposta) / teto * 100):.2f}%"
        ws[f"B{row_total + 5}"] = "MARGEM DISPONIVEL:"
        ws[f"F{row_total + 5}"] = round(teto - proposta, 2)


def _criar_breakeven(wb, postos, empresa_info):
    """Cria aba BREAK-EVEN com simulador de cenarios."""
    ws = wb.create_sheet("BREAK-EVEN")

    teto = empresa_info.get("valor_edital", 0)

    ws["A1"] = "ANALISE DE BREAK-EVEN E CENARIOS DE LANCE"
    ws["B3"] = "REFERENCIAS DO EDITAL"
    ws["B4"] = "Teto Edital (anual):"
    ws["E4"] = teto
    ws["B5"] = "Inexequibilidade (50%):"
    ws["E5"] = teto / 2 if teto else 0

    # Calcula custo puro (sem CI/Lucro)
    custo_puro_mensal = sum(
        p.get("custo_sem_margem", p.get("custo_total_mensal", 0)) * p.get("quantidade", 1)
        for p in postos
    )
    custo_puro_anual = custo_puro_mensal * 12

    ws["B7"] = "BREAK-EVEN POR CARGO (CI=0% / Lucro=0%)"

    cenarios = [
        ("Conservador (3%/3%)", 0.03, 0.03),
        ("Moderado (2.5%/2.5%)", 0.025, 0.025),
        ("Agressivo (2%/2%)", 0.02, 0.02),
        ("Muito agressivo (1.5%/1.5%)", 0.015, 0.015),
        ("Minimo (1%/1%)", 0.01, 0.01),
        ("Ultra-minimo (0.5%/0.5%)", 0.005, 0.005),
        ("BREAK-EVEN (0%/0%)", 0.0, 0.0),
    ]

    ws["B18"] = "SIMULADOR DE CENARIOS"
    ws["B19"] = "Cenario"
    ws["C19"] = "CI%"
    ws["D19"] = "Lucro%"
    ws["E19"] = "Valor Anual"
    ws["F19"] = "Desc/Teto"
    ws["G19"] = "Margem Bruta"
    ws["I19"] = "Status"

    for i, (nome, ci, lucro) in enumerate(cenarios):
        row = 20 + i
        # Recalcula com essa margem
        fator = (1 + ci + lucro)
        valor_anual = custo_puro_anual * fator
        desc = (teto - valor_anual) / teto * 100 if teto else 0
        margem = valor_anual - custo_puro_anual

        status = "CONFORTAVEL" if desc < 10 else "VIAVEL" if desc < 12 else "AGRESSIVO" if desc < 13 else "BREAK-EVEN"

        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = nome
        ws[f"C{row}"] = f"{ci*100:.1f}%"
        ws[f"D{row}"] = f"{lucro*100:.1f}%"
        ws[f"E{row}"] = round(valor_anual, 2)
        ws[f"F{row}"] = f"{desc:.2f}%"
        ws[f"G{row}"] = round(margem, 2)
        ws[f"I{row}"] = status


def pode_usar_template(template_path: Path) -> bool:
    """Verifica se o template e valido para preenchimento automatico."""
    try:
        wb = openpyxl.load_workbook(str(template_path), data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Verifica se tem estrutura de modulos
        has_modulo1 = False
        has_modulo6 = False
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = str(cell.value or "").lower()
                if "modulo 1" in v or "m\u00f3dulo 1" in v or "composi" in v:
                    has_modulo1 = True
                if "modulo 6" in v or "m\u00f3dulo 6" in v or "custos indiretos" in v:
                    has_modulo6 = True
        return has_modulo1 and has_modulo6
    except Exception:
        return False
