"""Gerador de planilha IN 05/2017 — 100% fórmulas Excel dinâmicas.

Cada posto gera 1 aba com Módulos 1-6.
Aba RESUMO consolida todos os postos.
Aba BREAK-EVEN simula 7 cenários de lance.

Todas as células de valor usam fórmulas Excel (=SUM, =cell*cell, etc.)
para que o usuário possa alterar inputs e ver recálculo automático.
"""
import logging
from pathlib import Path
from openpyxl import Workbook
import openpyxl.cell.cell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

log = logging.getLogger("planilha_builder")

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
_F = "Arial"
BOLD = Font(name=_F, bold=True, size=10)
BOLD_W = Font(name=_F, bold=True, size=10, color="FFFFFF")
BOLD_BLUE = Font(name=_F, bold=True, size=10, color="0000FF")
NORMAL = Font(name=_F, size=10)
SMALL = Font(name=_F, size=9, color="666666")
PCT_FONT = Font(name=_F, size=10, color="333333")
RED_BOLD = Font(name=_F, size=9, color="FF0000", bold=True)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
MOD_FILL = PatternFill("solid", fgColor="4472C4")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="CCCCCC")
BORDER_BOTTOM = Border(bottom=THIN)

FMT_BRL = '#,##0.00'
FMT_PCT = '0.00%'
ALIGN_CENTER = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_sheet_name(name: str) -> str:
    """Garante nome de aba <= 31 chars e sem caracteres inválidos."""
    for ch in "/\\*?[]:'":
        name = name.replace(ch, "-")
    return name.strip()[:31]


def _s(ws, row, col, value, font=NORMAL, fill=None, fmt=None, align=None):
    """Seta célula com estilo."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    return c


def _formula(ws, row, col, formula_str, font=NORMAL, fill=None, fmt=FMT_BRL):
    """Seta célula com fórmula Excel."""
    c = ws.cell(row=row, column=col, value=formula_str)
    c.font = font
    c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def _header_row(ws, row, col_b_text, col_c_text="%", col_d_text="Valor (R$)"):
    """Linha de subcabeçalho de módulo."""
    _s(ws, row, 2, col_b_text, BOLD)
    _s(ws, row, 3, col_c_text, BOLD, align=ALIGN_CENTER)
    _s(ws, row, 4, col_d_text, BOLD, align=ALIGN_CENTER)


def _mod_title(ws, row, text):
    """Título de módulo (mesclado, colorido)."""
    _s(ws, row, 1, text, BOLD_W, MOD_FILL)
    _s(ws, row, 2, "", BOLD_W, MOD_FILL)
    _s(ws, row, 3, "", BOLD_W, MOD_FILL)
    _s(ws, row, 4, "", BOLD_W, MOD_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def _total_row(ws, row, label, formula_str):
    """Linha de total de módulo."""
    _s(ws, row, 2, label, BOLD, TOTAL_FILL)
    _s(ws, row, 3, "", BOLD, TOTAL_FILL)
    _formula(ws, row, 4, formula_str, BOLD, TOTAL_FILL)


# ---------------------------------------------------------------------------
# Aba de posto (Módulos 1-6)
# ---------------------------------------------------------------------------
def _build_posto_sheet(ws, posto: dict, empresa_info: dict, licitacao_info: dict):
    """Constrói aba completa de um posto com Módulos 1-6 em fórmulas Excel.

    Layout fixo — as células-chave são:
        D9  = Salário base (input)
        D13 = Total Módulo 1
        D19 = Total 2.1 (13º/férias)
        D30 = Total 2.2 (GPS/FGTS)
        D38 = Total 2.3 (benefícios)
        D40 = Total Módulo 2
        D50 = Total Módulo 3
        D61 = Total Módulo 4
        D69 = Total Módulo 5
        D71 = Subtotal Módulos 1-5
        D75 = Custos Indiretos
        D76 = Lucro
        D81 = Total Tributos (por dentro)
        D82 = Total Módulo 6
        D78 = Valor Mensal por Empregado  (*** referência do RESUMO ***)
    """
    # --- Extrai dados do posto ---
    funcao = posto.get("funcao", "CARGO")
    qtd = posto.get("quantidade", 1) or 1
    jornada = posto.get("jornada", "44h")
    salario = posto.get("salario_base", 0)

    # Adicionais (percentuais inteiros -> fração)
    insalub_pct = (posto.get("adicional_insalubridade_pct", 0) or 0) / 100
    peric_pct = (posto.get("adicional_periculosidade_pct", 0) or 0) / 100
    noturno = posto.get("adicional_noturno", False)
    # Adicional noturno: 20% sobre salário. Para 12x36, proporcional ~50% horas noturnas.
    noturno_pct = 0.0
    if noturno:
        noturno_pct = 0.20 * 0.5 if jornada == "12x36" else 0.20

    # Benefícios
    vt = posto.get("vt", 0) or 0
    va = posto.get("va", 0) or 0
    cesta = posto.get("cesta_basica", 0) or 0
    bsf = posto.get("bsf", 0) or 0
    seguro = posto.get("seguro_vida", 0) or 0

    # Insumos
    uniformes = posto.get("uniformes", 85) if posto.get("uniformes") is not None else 85
    materiais = posto.get("materiais", 0) or 0
    equipamentos = posto.get("equipamentos", 0) or 0

    # Empresa / encargos
    sat_rat = empresa_info.get("sat_rat_pct", 0.03)
    ci_pct = empresa_info.get("ci_pct", 0.03)
    lucro_pct = empresa_info.get("lucro_pct", 0.03)
    pis_pct = empresa_info.get("pis_pct", 0.0005)
    cofins_pct = empresa_info.get("cofins_pct", 0.0415)
    iss_pct = empresa_info.get("iss_pct", 0.02)

    cct_nome = licitacao_info.get("cct_nome", "")

    # --- Larguras ---
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18

    # =====================================================================
    # CABEÇALHO  (linhas 1-6)
    # =====================================================================
    _s(ws, 1, 1, f"PLANILHA DE CUSTOS E FORMAÇÃO DE PREÇOS", BOLD)
    ws.merge_cells("A1:D1")
    _s(ws, 2, 1, "CCT:")
    _s(ws, 2, 2, cct_nome, BOLD_BLUE, INPUT_FILL)
    _s(ws, 3, 1, "Cargo:")
    _s(ws, 3, 2, funcao.upper(), BOLD)
    _s(ws, 3, 3, "Jornada:")
    _s(ws, 3, 4, jornada)
    _s(ws, 4, 1, "Postos:")
    _s(ws, 4, 2, qtd, BOLD_BLUE, INPUT_FILL)  # B4 = quantidade
    _s(ws, 5, 1, "Salário:")
    _s(ws, 5, 2, salario, BOLD_BLUE, INPUT_FILL, FMT_BRL)

    # =====================================================================
    # MÓDULO 1 — Composição da Remuneração  (linhas 7-13)
    # =====================================================================
    _mod_title(ws, 7, "MÓDULO 1 - COMPOSIÇÃO DA REMUNERAÇÃO")
    _header_row(ws, 8, "Composição da Remuneração")

    # D9: Salário base (INPUT)
    _s(ws, 9, 1, "A")
    _s(ws, 9, 2, "Salário Base")
    _s(ws, 9, 4, salario, BOLD_BLUE, INPUT_FILL, FMT_BRL)

    # D10: Adicional de periculosidade
    _s(ws, 10, 1, "B")
    _s(ws, 10, 2, "Adicional de periculosidade")
    _s(ws, 10, 3, peric_pct, PCT_FONT, None, FMT_PCT)
    if peric_pct:
        _formula(ws, 10, 4, "=C10*D9")
    else:
        _s(ws, 10, 4, 0, NORMAL, None, FMT_BRL)

    # D11: Adicional de insalubridade
    _s(ws, 11, 1, "C")
    _s(ws, 11, 2, "Adicional de insalubridade")
    _s(ws, 11, 3, insalub_pct, PCT_FONT, None, FMT_PCT)
    if insalub_pct:
        _formula(ws, 11, 4, "=C11*D9")
    else:
        _s(ws, 11, 4, 0, NORMAL, None, FMT_BRL)

    # D12: Adicional noturno
    _s(ws, 12, 1, "D")
    _s(ws, 12, 2, "Adicional noturno")
    _s(ws, 12, 3, noturno_pct, PCT_FONT, None, FMT_PCT)
    if noturno_pct:
        _formula(ws, 12, 4, "=C12*D9")
    else:
        _s(ws, 12, 4, 0, NORMAL, None, FMT_BRL)

    # D13: TOTAL MÓDULO 1
    _total_row(ws, 13, "TOTAL MÓDULO 1", "=SUM(D9:D12)")

    # =====================================================================
    # MÓDULO 2 — Encargos e Benefícios  (linhas 15-40)
    # =====================================================================
    _mod_title(ws, 15, "MÓDULO 2 - ENCARGOS E BENEFÍCIOS")

    # --- Submódulo 2.1: 13º e Férias (linhas 16-19) ---
    _s(ws, 16, 1, "2.1")
    _header_row(ws, 16, "13º Salário, Férias e Adicional de Férias")

    _s(ws, 17, 1, "A")
    _s(ws, 17, 2, "13º Salário (1/12)")
    _s(ws, 17, 3, 1/12, PCT_FONT, None, FMT_PCT)
    _formula(ws, 17, 4, "=D13/12")

    _s(ws, 18, 1, "B")
    _s(ws, 18, 2, "Férias + 1/3 constitucional (1/12 x 4/3)")
    _s(ws, 18, 3, 4/36, PCT_FONT, None, FMT_PCT)
    _formula(ws, 18, 4, "=D13*(4/3)/12")

    # D19: Total 2.1
    _s(ws, 19, 2, "Total Submódulo 2.1", BOLD)
    _formula(ws, 19, 4, "=SUM(D17:D18)", BOLD)

    # --- Submódulo 2.2: GPS, FGTS e Contribuições (linhas 21-30) ---
    # Base = Módulo 1 + Submódulo 2.1  =>  (D13+D19)
    _s(ws, 21, 1, "2.2")
    _header_row(ws, 21, "GPS, FGTS e Outras Contribuições")

    encargos_22 = [
        ("A", "INSS Patronal", 0.20),
        ("B", "Salário Educação", 0.025),
        ("C", "SAT/RAT", sat_rat),
        ("D", "SESC ou SESI", 0.015),
        ("E", "SENAC ou SENAI", 0.01),
        ("F", "SEBRAE", 0.006),
        ("G", "INCRA", 0.002),
        ("H", "FGTS", 0.08),
    ]
    for i, (letra, nome_enc, pct) in enumerate(encargos_22):
        row = 22 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, nome_enc)
        _s(ws, row, 3, pct, PCT_FONT, None, FMT_PCT)
        # Base de cálculo = D13 + D19 (Mód1 + Sub2.1)
        _formula(ws, row, 4, f"=C{row}*(D13+D19)")

    # D30: Total 2.2
    _s(ws, 30, 2, "Total Submódulo 2.2", BOLD)
    _formula(ws, 30, 4, "=SUM(D22:D29)", BOLD)

    # --- Submódulo 2.3: Benefícios Mensais (linhas 32-38) ---
    _s(ws, 32, 1, "2.3")
    _header_row(ws, 32, "Benefícios Mensais e Diários", "", "Valor (R$)")

    beneficios = [
        ("A", "Vale-Transporte (líquido)", vt),
        ("B", "Auxílio-Alimentação/Refeição", va),
        ("C", "Cesta Básica", cesta),
        ("D", "Benefício Social Familiar (BSF)", bsf),
        ("E", "Seguro de Vida", seguro),
    ]
    for i, (letra, nome_b, valor) in enumerate(beneficios):
        row = 33 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, nome_b)
        _s(ws, row, 4, valor, BOLD_BLUE, INPUT_FILL, FMT_BRL)

    # D38: Total 2.3
    _s(ws, 38, 2, "Total Submódulo 2.3", BOLD)
    _formula(ws, 38, 4, "=SUM(D33:D37)", BOLD)

    # D40: TOTAL MÓDULO 2
    _total_row(ws, 40, "TOTAL MÓDULO 2", "=D19+D30+D38")

    # =====================================================================
    # MÓDULO 3 — Provisão para Rescisão  (linhas 42-50)
    # Base = D13 (Módulo 1)
    # =====================================================================
    _mod_title(ws, 42, "MÓDULO 3 - PROVISÃO PARA RESCISÃO")
    _s(ws, 43, 1, "3")
    _header_row(ws, 43, "Provisão para Rescisão")

    rescisao = [
        ("A", "Aviso prévio indenizado", 0.0417),
        ("B", "Incidência FGTS sobre API", 0.0033),
        ("C", "Multa FGTS + CS sobre API", 0.0167),
        ("D", "Aviso prévio trabalhado", 0.0194),
        ("E", "Incidência GPS/FGTS sobre APT", 0.0071),
        ("F", "Multa FGTS + CS sobre APT", 0.0078),
    ]
    for i, (letra, nome_r, pct) in enumerate(rescisao):
        row = 44 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, nome_r)
        _s(ws, row, 3, pct, PCT_FONT, None, FMT_PCT)
        _formula(ws, row, 4, f"=C{row}*D13")

    # D50: TOTAL MÓDULO 3
    _total_row(ws, 50, "TOTAL MÓDULO 3", "=SUM(D44:D49)")

    # =====================================================================
    # MÓDULO 4 — Custo de Reposição do Profissional Ausente  (linhas 52-61)
    # Base = D13 + D40  (Módulo 1 + Módulo 2)
    # =====================================================================
    _mod_title(ws, 52, "MÓDULO 4 - REPOSIÇÃO DO PROFISSIONAL AUSENTE")
    _s(ws, 53, 1, "4.1")
    _header_row(ws, 53, "Ausências Legais")

    ausencias = [
        ("A", "Férias (substituto)", 0.0833),
        ("B", "Ausências legais", 0.0166),
        ("C", "Licença-paternidade", 0.0014),
        ("D", "Acidente de trabalho", 0.0083),
        ("E", "Afastamento maternidade", 0.0033),
    ]
    for i, (letra, nome_a, pct) in enumerate(ausencias):
        row = 54 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, nome_a)
        _s(ws, row, 3, pct, PCT_FONT, None, FMT_PCT)
        # Base = Módulo 1 + Módulo 2 total
        _formula(ws, row, 4, f"=C{row}*(D13+D40)")

    # D59: Total 4.1
    _s(ws, 59, 2, "Total Submódulo 4.1", BOLD)
    _formula(ws, 59, 4, "=SUM(D54:D58)", BOLD)

    # 4.2 Intrajornada
    _s(ws, 60, 1, "4.2")
    _s(ws, 60, 2, "Intrajornada")
    _s(ws, 60, 4, 0, NORMAL, None, FMT_BRL)

    # D61: TOTAL MÓDULO 4
    _total_row(ws, 61, "TOTAL MÓDULO 4", "=D59+D60")

    # =====================================================================
    # MÓDULO 5 — Insumos Diversos  (linhas 63-69)
    # =====================================================================
    _mod_title(ws, 63, "MÓDULO 5 - INSUMOS DIVERSOS")
    _s(ws, 64, 1, "5")
    _header_row(ws, 64, "Insumos Diversos", "", "Valor (R$)")

    insumos = [
        ("A", "Uniformes (mensal)", uniformes),
        ("B", "Materiais", materiais),
        ("C", "Equipamentos", equipamentos),
    ]
    for i, (letra, nome_i, valor) in enumerate(insumos):
        row = 65 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, nome_i)
        _s(ws, row, 4, valor, BOLD_BLUE, INPUT_FILL, FMT_BRL)

    # D69: TOTAL MÓDULO 5  (linhas 65-68 para ter espaço para D futuramente)
    # Usamos 65:67 pois só temos 3 linhas, mas mantemos D69 no lugar fixo
    _s(ws, 68, 4, 0, NORMAL, None, FMT_BRL)  # placeholder linha D vazia
    _total_row(ws, 69, "TOTAL MÓDULO 5", "=SUM(D65:D68)")

    # =====================================================================
    # SUBTOTAL Módulos 1-5  (linha 71)
    # =====================================================================
    _s(ws, 71, 2, "SUBTOTAL (Módulos 1 a 5)", BOLD, SUBTOTAL_FILL)
    _s(ws, 71, 3, "", BOLD, SUBTOTAL_FILL)
    _formula(ws, 71, 4, "=D13+D40+D50+D61+D69", BOLD, SUBTOTAL_FILL)

    # =====================================================================
    # MÓDULO 6 — Custos Indiretos, Tributos e Lucro  (linhas 73-82)
    # Base = D71 (subtotal M1-M5)
    # Tributos POR DENTRO: =(D71+D75+D76)*C81/(1-C81)
    # =====================================================================
    _mod_title(ws, 73, "MÓDULO 6 - CUSTOS INDIRETOS, TRIBUTOS E LUCRO")
    _s(ws, 74, 1, "6")
    _header_row(ws, 74, "Custos Indiretos, Tributos e Lucro")

    # D75: Custos Indiretos
    _s(ws, 75, 1, "A")
    _s(ws, 75, 2, "Custos Indiretos")
    _s(ws, 75, 3, ci_pct, BOLD_BLUE, INPUT_FILL, FMT_PCT)
    _formula(ws, 75, 4, "=C75*D71")

    # D76: Lucro
    _s(ws, 76, 1, "B")
    _s(ws, 76, 2, "Lucro")
    _s(ws, 76, 3, lucro_pct, BOLD_BLUE, INPUT_FILL, FMT_PCT)
    _formula(ws, 76, 4, "=C76*D71")

    # Tributos (detalhamento)
    _s(ws, 77, 1, "C")
    _s(ws, 77, 2, "Tributos", BOLD)
    _s(ws, 78, 2, "  C.1 PIS")
    _s(ws, 78, 3, pis_pct, PCT_FONT, None, FMT_PCT)
    _s(ws, 79, 2, "  C.2 COFINS")
    _s(ws, 79, 3, cofins_pct, PCT_FONT, None, FMT_PCT)
    _s(ws, 80, 2, "  C.3 ISS")
    _s(ws, 80, 3, iss_pct, PCT_FONT, None, FMT_PCT)

    # C81: Total tributos %
    _s(ws, 81, 2, "  Total Tributos %")
    _formula(ws, 81, 3, "=SUM(C78:C80)", BOLD, fmt=FMT_PCT)
    # D81: Tributos POR DENTRO  =>  =(D71+D75+D76)*C81/(1-C81)
    _formula(ws, 81, 4, "=(D71+D75+D76)*C81/(1-C81)")

    # D82: TOTAL MÓDULO 6
    _total_row(ws, 82, "TOTAL MÓDULO 6", "=D75+D76+D81")

    # =====================================================================
    # QUADRO-RESUMO  (linhas 84-94)
    # D78 é referenciada pelo RESUMO como "valor mensal por empregado"
    # Porém pela especificação, D78 deve conter esse valor.
    # Vamos colocar o VALOR MENSAL POR EMPREGADO em D78 da seção de resumo
    # abaixo, e ajustar o layout para que D78 = valor mensal/empregado.
    # =====================================================================
    # ATENÇÃO: A especificação pede que o RESUMO referencie D78 de cada aba.
    # Vamos usar a linha 78 como valor-chave. Porém acima D78 já tem C.1 PIS.
    # Solução: movemos o quadro-resumo para que a linha-chave D78 fique
    # disponível. Recalculando o layout:
    #   - Tributos C.1 PIS vai para outra posição
    # ALTERNATIVA: manter o layout e colocar o VALOR MENSAL em D78 usando
    # uma fórmula extra. Vamos optar por NÃO alterar o layout dos módulos
    # e simplesmente referenciar D92 no RESUMO como "valor mensal".
    # MAS a especificação diz explicitamente "D78". Então vamos reorganizar.

    # --- REORGANIZAÇÃO para ter D78 = VALOR MENSAL POR EMPREGADO ---
    # Precisamos que a linha 78 seja o valor final.
    # Vamos comprimir o Módulo 6 para que tudo caiba antes da linha 78.

    # RECONSTRUINDO: vamos limpar o que fizemos acima do módulo 6 e refazer
    # com layout que termina o quadro-resumo na linha 78.

    # Na verdade, o mais limpo é simplesmente inserir uma célula D78 com a
    # fórmula do valor mensal, e mover os tributos para não conflitar.
    # Mas D78 já está ocupada com PIS.
    # Precisamos de um layout diferente. Vamos refazer todo o módulo 6.

    # ===> DECISÃO: Reescrever módulo 6 em layout compacto para liberar D78.

    # Limpar células do módulo 6 (evitar MergedCell errors)
    for row in range(73, 83):
        for col in range(1, 5):
            try:
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None
                    cell.font = NORMAL
                    cell.fill = PatternFill()
                    cell.number_format = 'General'
            except (AttributeError, TypeError):
                pass
    # Unmerge any merged ranges in this area
    ranges_to_remove = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= 73 and mr.max_row <= 82:
            ranges_to_remove.append(str(mr))
    for mr_str in ranges_to_remove:
        ws.unmerge_cells(mr_str)

    # === MÓDULO 6 REFEITO — layout compacto (linhas 73-77) ===
    _mod_title(ws, 73, "MÓDULO 6 - CUSTOS INDIRETOS, TRIBUTOS E LUCRO")

    # Linha 74: CI
    _s(ws, 74, 1, "A")
    _s(ws, 74, 2, "Custos Indiretos")
    _s(ws, 74, 3, ci_pct, BOLD_BLUE, INPUT_FILL, FMT_PCT)
    _formula(ws, 74, 4, "=C74*D71")

    # Linha 75: Lucro
    _s(ws, 75, 1, "B")
    _s(ws, 75, 2, "Lucro")
    _s(ws, 75, 3, lucro_pct, BOLD_BLUE, INPUT_FILL, FMT_PCT)
    _formula(ws, 75, 4, "=C75*D71")

    # Linha 76: Total tributos % (PIS+COFINS+ISS como input direto na célula)
    trib_total = pis_pct + cofins_pct + iss_pct
    _s(ws, 76, 1, "C")
    _s(ws, 76, 2, f"Tributos (PIS {pis_pct*100:.2f}% + COFINS {cofins_pct*100:.2f}% + ISS {iss_pct*100:.1f}%)")
    _s(ws, 76, 3, trib_total, PCT_FONT, INPUT_FILL, FMT_PCT)
    # Tributos POR DENTRO: =(D71+D74+D75)*C76/(1-C76)
    _formula(ws, 76, 4, "=(D71+D74+D75)*C76/(1-C76)")

    # Linha 77: TOTAL MÓDULO 6
    _total_row(ws, 77, "TOTAL MÓDULO 6", "=D74+D75+D76")

    # === D78: VALOR MENSAL POR EMPREGADO === (referência do RESUMO)
    _s(ws, 78, 1, "", BOLD_W, HEADER_FILL)
    _s(ws, 78, 2, "VALOR MENSAL POR EMPREGADO", BOLD_W, HEADER_FILL)
    _s(ws, 78, 3, "", BOLD_W, HEADER_FILL)
    _formula(ws, 78, 4, "=D71+D77", BOLD_W, HEADER_FILL)
    ws.merge_cells("A78:C78")

    # === QUADRO-RESUMO (linhas 80-94) ===
    _s(ws, 80, 1, "QUADRO-RESUMO DO CUSTO POR EMPREGADO", BOLD_W, HEADER_FILL)
    _s(ws, 80, 2, "", BOLD_W, HEADER_FILL)
    _s(ws, 80, 3, "", BOLD_W, HEADER_FILL)
    _s(ws, 80, 4, "", BOLD_W, HEADER_FILL)
    ws.merge_cells("A80:D80")

    resumo_items = [
        ("A", "Módulo 1 - Composição da Remuneração", "=D13"),
        ("B", "Módulo 2 - Encargos e Benefícios", "=D40"),
        ("C", "Módulo 3 - Provisão para Rescisão", "=D50"),
        ("D", "Módulo 4 - Custo de Reposição", "=D61"),
        ("E", "Módulo 5 - Insumos Diversos", "=D69"),
    ]
    for i, (letra, desc, ref) in enumerate(resumo_items):
        row = 81 + i
        _s(ws, row, 1, letra)
        _s(ws, row, 2, desc)
        _formula(ws, row, 4, ref)

    # Subtotal M1-M5
    _s(ws, 86, 2, "SUBTOTAL (A+B+C+D+E)", BOLD, TOTAL_FILL)
    _s(ws, 86, 3, "", BOLD, TOTAL_FILL)
    _formula(ws, 86, 4, "=D71", BOLD, TOTAL_FILL)

    _s(ws, 87, 1, "F")
    _s(ws, 87, 2, "Módulo 6 - CI, Tributos e Lucro")
    _formula(ws, 87, 4, "=D77")

    # Valor mensal / empregado (repete D78)
    _s(ws, 88, 2, "VALOR MENSAL POR EMPREGADO", BOLD_W, HEADER_FILL)
    _s(ws, 88, 3, "", BOLD_W, HEADER_FILL)
    _formula(ws, 88, 4, "=D78", BOLD_W, HEADER_FILL)

    # Valor mensal (postos)
    _s(ws, 89, 2, f"VALOR MENSAL ({qtd} postos)", BOLD)
    _formula(ws, 89, 4, "=D78*B4", BOLD)

    # Valor anual
    _s(ws, 90, 2, f"VALOR ANUAL ({qtd} postos x 12)", BOLD)
    _formula(ws, 90, 4, "=D89*12", BOLD)


# ---------------------------------------------------------------------------
# Aba RESUMO
# ---------------------------------------------------------------------------
def _build_resumo_sheet(ws, postos, sheet_names, empresa_info, licitacao_info):
    """Constrói aba RESUMO consolidando todos os postos.

    Referencia D78 de cada aba de posto (valor mensal por empregado).
    """
    orgao = licitacao_info.get("orgao", "")
    objeto = (licitacao_info.get("objeto", "") or "")[:80]
    empresa = empresa_info.get("nome", "")
    regime = empresa_info.get("regime", "")
    prazo = licitacao_info.get("prazo_meses", 12)
    valor_teto = licitacao_info.get("valor_teto", 0) or 0
    pregao = licitacao_info.get("pregao", "")

    pis = empresa_info.get("pis_pct", 0.0005)
    cofins = empresa_info.get("cofins_pct", 0.0415)
    iss = empresa_info.get("iss_pct", 0.02)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Cabeçalho
    _s(ws, 1, 1, f"RESUMO - {orgao}", BOLD)
    ws.merge_cells("A1:F1")
    _s(ws, 2, 1, f"Pregão: {pregao} | Objeto: {objeto}", SMALL)
    ws.merge_cells("A2:F2")
    _s(ws, 3, 1, f"Empresa: {empresa} | Regime: {regime} | PIS {pis*100:.2f}% COFINS {cofins*100:.2f}% ISS {iss*100:.1f}%", SMALL)
    ws.merge_cells("A3:F3")

    # Cabeçalho tabela
    r = 5
    for col, header in enumerate(["Item", "Cargo", "Postos", "R$/Emp/Mês", "R$ Mensal", "R$ Anual"], 1):
        _s(ws, r, col, header, BOLD_W, HEADER_FILL, align=ALIGN_CENTER)

    # Linhas de postos
    for i, (posto, sname) in enumerate(zip(postos, sheet_names)):
        row = 6 + i
        _s(ws, row, 1, i + 1, align=ALIGN_CENTER)
        _s(ws, row, 2, posto.get("funcao", "").upper(), BOLD)
        # Quantidade = B4 da aba
        _formula(ws, row, 3, f"='{sname}'!B4", fmt='0')
        # Valor por empregado/mês = D78 da aba
        _formula(ws, row, 4, f"='{sname}'!D78")
        # Mensal = postos x valor unitário
        _formula(ws, row, 5, f"=C{row}*D{row}")
        # Anual = mensal x 12
        _formula(ws, row, 6, f"=E{row}*12")

    # Total
    rt = 6 + len(postos)
    _s(ws, rt, 2, "TOTAL", BOLD, TOTAL_FILL)
    _formula(ws, rt, 3, f"=SUM(C6:C{rt-1})", BOLD, TOTAL_FILL, fmt='0')
    _formula(ws, rt, 5, f"=SUM(E6:E{rt-1})", BOLD, TOTAL_FILL)
    _formula(ws, rt, 6, f"=SUM(F6:F{rt-1})", BOLD, TOTAL_FILL)

    # Comparação com teto
    r = rt + 2
    _s(ws, r, 2, "TETO EDITAL (anual):", BOLD)
    _s(ws, r, 6, valor_teto, BOLD, INPUT_FILL, FMT_BRL)
    teto_r = r
    r += 1
    _s(ws, r, 2, "NOSSA PROPOSTA (anual):", BOLD)
    _formula(ws, r, 6, f"=F{rt}", BOLD)
    prop_r = r
    r += 1
    _s(ws, r, 2, "DESCONTO (%):", BOLD)
    _formula(ws, r, 6, f"=IF(F{teto_r}=0,0,1-F{prop_r}/F{teto_r})", BOLD, fmt=FMT_PCT)
    r += 1
    _s(ws, r, 2, "MARGEM DISPONÍVEL (R$):", BOLD)
    _formula(ws, r, 6, f"=F{teto_r}-F{prop_r}", BOLD)
    r += 2
    _s(ws, r, 2, f"VALOR CONTRATO ({prazo} meses):", BOLD)
    _formula(ws, r, 6, f"=E{rt}*{prazo}", BOLD)


# ---------------------------------------------------------------------------
# Aba BREAK-EVEN
# ---------------------------------------------------------------------------
def _build_breakeven_sheet(ws, postos, sheet_names, empresa_info, licitacao_info):
    """Constrói aba BREAK-EVEN com 7 cenários variando CI/Lucro."""
    valor_teto = licitacao_info.get("valor_teto", 0) or 0

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16

    _s(ws, 1, 1, "ANÁLISE DE BREAK-EVEN E CENÁRIOS DE LANCE", BOLD)
    ws.merge_cells("A1:I1")

    # Referências
    _s(ws, 3, 2, "REFERÊNCIAS DO EDITAL", BOLD)
    _s(ws, 4, 2, "Teto Edital (anual):")
    _s(ws, 4, 5, valor_teto, BOLD, INPUT_FILL, FMT_BRL)  # E4
    _s(ws, 5, 2, "Inexequibilidade (50%):")
    _formula(ws, 5, 5, "=E4*0.5")

    # Break-even por cargo (custo puro = D71 de cada aba = subtotal sem M6)
    _s(ws, 7, 2, "BREAK-EVEN POR CARGO (CI=0% / Lucro=0%)", BOLD)
    for col, h in [(2, "Cargo"), (3, "Postos"), (5, "Custo/Emp/Mês"), (6, "Custo Mensal"), (9, "Custo Anual")]:
        _s(ws, 8, col, h, BOLD_W, HEADER_FILL, align=ALIGN_CENTER)

    for i, (posto, sname) in enumerate(zip(postos, sheet_names)):
        row = 9 + i
        _s(ws, row, 1, i + 1, align=ALIGN_CENTER)
        _s(ws, row, 2, posto.get("funcao", "").upper())
        _formula(ws, row, 3, f"='{sname}'!B4", fmt='0')
        # Custo puro = D71 (subtotal M1-M5, sem CI/Lucro/Tributos)
        _formula(ws, row, 5, f"='{sname}'!D71")
        _formula(ws, row, 6, f"=C{row}*E{row}")
        _formula(ws, row, 9, f"=F{row}*12")

    r_be = 9 + len(postos)
    _s(ws, r_be, 2, "CUSTO PURO (BREAK-EVEN)", BOLD, TOTAL_FILL)
    _formula(ws, r_be, 6, f"=SUM(F9:F{r_be-1})", BOLD, TOTAL_FILL)
    _formula(ws, r_be, 9, f"=SUM(I9:I{r_be-1})", BOLD, TOTAL_FILL)
    _s(ws, r_be + 1, 2, "Abaixo deste valor = PREJUÍZO", RED_BOLD)

    # Total tributos % — referência da 1ª aba de posto (C76)
    trib_ref = f"'{sheet_names[0]}'!C76" if sheet_names else "0"

    # Cenários
    r = r_be + 3
    _s(ws, r, 2, "SIMULADOR DE CENÁRIOS", BOLD)
    r += 1
    for col, h in [(2, "Cenário"), (3, "CI%"), (4, "Lucro%"), (5, "Valor Anual"),
                   (6, "Desc/Teto"), (7, "Margem Bruta"), (8, "Margem/Mês"), (9, "Status")]:
        _s(ws, r, col, h, BOLD_W, HEADER_FILL, align=ALIGN_CENTER)

    cenarios = [
        ("Conservador (3%/3%)", 0.03, 0.03),
        ("Moderado (2.5%/2.5%)", 0.025, 0.025),
        ("Agressivo (2%/2%)", 0.02, 0.02),
        ("Muito agressivo (1.5%/1.5%)", 0.015, 0.015),
        ("Mínimo (1%/1%)", 0.01, 0.01),
        ("Ultra-mínimo (0.5%/0.5%)", 0.005, 0.005),
        ("BREAK-EVEN (0%/0%)", 0.0, 0.0),
    ]

    r_header = r
    for i, (nome_c, ci, lucro) in enumerate(cenarios):
        cr = r_header + 1 + i
        _s(ws, cr, 1, i + 1, align=ALIGN_CENTER)
        _s(ws, cr, 2, nome_c)
        _s(ws, cr, 3, ci, None, None, FMT_PCT)
        _s(ws, cr, 4, lucro, None, None, FMT_PCT)
        # Valor Anual = custo_puro_anual * (1+CI+Lucro) / (1-trib%)
        be_cell = f"I{r_be}"
        _formula(ws, cr, 5, f"={be_cell}*(1+C{cr}+D{cr})/(1-{trib_ref})")
        # Desconto / Teto
        _formula(ws, cr, 6, f"=IF(E4=0,0,1-E{cr}/E4)", fmt=FMT_PCT)
        # Margem Bruta = Valor - custo_puro/(1-trib)
        _formula(ws, cr, 7, f"=E{cr}-{be_cell}/(1-{trib_ref})")
        # Margem/Mês
        _formula(ws, cr, 8, f"=G{cr}/12")
        # Status
        ws.cell(row=cr, column=9).value = (
            f'=IF(E{cr}<E5,"INEXEQUÍVEL",'
            f'IF(E{cr}>E4,"ACIMA REF",'
            f'IF(F{cr}>0.15,"RISCO","OK")))'
        )

    # Resumo estratégico
    r_first = r_header + 1
    r = r_first + len(cenarios) + 1
    _s(ws, r, 2, "RESUMO ESTRATÉGICO", BOLD)
    r += 1
    _s(ws, r, 2, "Lance inicial (3%/3%):", BOLD)
    _formula(ws, r, 5, f"=E{r_first}")
    _formula(ws, r, 6, f"=F{r_first}", fmt=FMT_PCT)
    r += 1
    _s(ws, r, 2, "Lance competitivo (2%/2%):", BOLD)
    _formula(ws, r, 5, f"=E{r_first+2}")
    _formula(ws, r, 6, f"=F{r_first+2}", fmt=FMT_PCT)
    r += 1
    _s(ws, r, 2, "PISO ABSOLUTO (break-even):", BOLD)
    _formula(ws, r, 5, f"=E{r_first+6}")
    _formula(ws, r, 6, f"=F{r_first+6}", fmt=FMT_PCT)
    r += 1
    _s(ws, r, 2, "Inexequibilidade (50% teto):")
    _formula(ws, r, 5, "=E5")
    r += 2
    _s(ws, r, 2, "DESCONTO MÁXIMO SUSTENTÁVEL:", BOLD)
    _formula(ws, r, 5, f"=F{r_first+6}", fmt=FMT_PCT)
    r += 1
    _s(ws, r, 2, "VALOR MÍNIMO ABSOLUTO (anual):", BOLD)
    _formula(ws, r, 5, f"=E{r_first+6}")
    r += 1
    _s(ws, r, 2, "VALOR MÍNIMO ABSOLUTO (mensal):", BOLD)
    _formula(ws, r, 5, f"=E{r_first+6}/12")


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------
def gerar_planilha(
    postos: list[dict],
    empresa_info: dict,
    licitacao_info: dict,
    output_path: str | Path,
    incluir_breakeven: bool = True,
) -> Path:
    """Gera planilha IN 05/2017 completa com fórmulas Excel dinâmicas.

    Args:
        postos: Lista de dicts com dados de cada posto. Cada dict contém:
            - funcao (str): nome do cargo
            - quantidade (int): número de postos
            - jornada (str): "44h", "30h", "12x36", etc.
            - salario_base (float): salário base mensal
            - adicional_insalubridade_pct (int): 0, 20 ou 40
            - adicional_periculosidade_pct (int): 0 ou 30
            - adicional_noturno (bool): se tem adicional noturno
            - vt (float): vale-transporte líquido
            - va (float): vale-alimentação
            - cesta_basica (float): default 0
            - bsf (float): benefício social familiar, default 0
            - seguro_vida (float): default 0
            - uniformes (float): default 85
            - materiais (float): default 0
            - equipamentos (float): default 0

        empresa_info: Dict com dados da empresa:
            - nome, regime, ci_pct, lucro_pct
            - pis_pct, cofins_pct, iss_pct, sat_rat_pct

        licitacao_info: Dict com dados da licitação:
            - orgao, pregao, objeto, valor_teto
            - cct_nome, prazo_meses

        output_path: Caminho do arquivo .xlsx de saída

    Returns:
        Path do arquivo gerado
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = []

    # 1. Uma aba por posto
    for posto in postos:
        funcao = posto.get("funcao", "CARGO")
        sname = _safe_sheet_name(funcao.upper())
        # Evitar duplicatas
        base = sname
        counter = 2
        while sname in sheet_names:
            sname = _safe_sheet_name(f"{base}_{counter}")
            counter += 1
        sheet_names.append(sname)

        ws = wb.create_sheet(title=sname)
        _build_posto_sheet(ws, posto, empresa_info, licitacao_info)

    # 2. Aba RESUMO
    ws_resumo = wb.create_sheet(title="RESUMO")
    _build_resumo_sheet(ws_resumo, postos, sheet_names, empresa_info, licitacao_info)

    # 3. Aba BREAK-EVEN (opcional — omitida em planilhas para entrega ao órgão)
    total_abas = len(sheet_names) + 1
    if incluir_breakeven:
        ws_be = wb.create_sheet(title="BREAK-EVEN")
        _build_breakeven_sheet(ws_be, postos, sheet_names, empresa_info, licitacao_info)
        total_abas += 1

    wb.save(str(output_path))
    log.info(f"Planilha gerada: {output_path} ({len(postos)} postos, {total_abas} abas, "
             f"{'com' if incluir_breakeven else 'SEM'} break-even)")
    return output_path
