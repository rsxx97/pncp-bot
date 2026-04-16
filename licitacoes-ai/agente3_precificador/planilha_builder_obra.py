"""Gerador de planilha orçamentária para OBRAS e REFORMAS.

Modelo diferente da IN 05/2017 (que é para mão de obra terceirizada).
Aqui: itens de serviço de engenharia + BDI de obra (referência TCU).

Todas as células de valor usam fórmulas Excel dinâmicas.
"""
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

log = logging.getLogger("planilha_builder_obra")

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
_F = "Arial"
BOLD = Font(name=_F, bold=True, size=10)
BOLD_W = Font(name=_F, bold=True, size=10, color="FFFFFF")
BOLD_BLUE = Font(name=_F, bold=True, size=10, color="0000FF")
NORMAL = Font(name=_F, size=10)
SMALL = Font(name=_F, size=9, color="666666")
RED_BOLD = Font(name=_F, size=9, color="FF0000", bold=True)
TITLE_FONT = Font(name=_F, bold=True, size=12, color="FFFFFF")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
MOD_FILL = PatternFill("solid", fgColor="4472C4")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
ALERT_FILL = PatternFill("solid", fgColor="FCE4EC")

THIN = Side(style="thin", color="CCCCCC")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FMT_BRL = '#,##0.00'
FMT_PCT = '0.00%'
FMT_INT = '#,##0'
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def _s(ws, row, col, value, font=NORMAL, fill=None, fmt=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = BORDER_ALL
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    return c


def _formula(ws, row, col, formula_str, font=NORMAL, fill=None, fmt=FMT_BRL):
    c = ws.cell(row=row, column=col, value=formula_str)
    c.font = font
    c.number_format = fmt
    c.border = BORDER_ALL
    if fill:
        c.fill = fill
    return c


# ---------------------------------------------------------------------------
# Itens padrão por categoria de obra
# ---------------------------------------------------------------------------
ITENS_OBRA = {
    "construcao": [
        {"item": "Serviços Preliminares", "descricao": "Instalação de canteiro, tapumes, placa de obra", "unid": "vb", "peso_pct": 3.0},
        {"item": "Fundações", "descricao": "Escavação, estacas, blocos de fundação, vigas baldrame", "unid": "vb", "peso_pct": 10.0},
        {"item": "Estrutura", "descricao": "Concreto armado, formas, armação, pilares, vigas, lajes", "unid": "vb", "peso_pct": 20.0},
        {"item": "Alvenaria e Vedação", "descricao": "Alvenaria de bloco cerâmico/concreto, vergas, contravergas", "unid": "vb", "peso_pct": 8.0},
        {"item": "Cobertura", "descricao": "Estrutura metálica/madeira, telhas, calhas, rufos", "unid": "vb", "peso_pct": 6.0},
        {"item": "Instalações Hidráulicas", "descricao": "Água fria/quente, esgoto, águas pluviais, louças/metais", "unid": "vb", "peso_pct": 8.0},
        {"item": "Instalações Elétricas", "descricao": "Quadros, fiação, tomadas, iluminação, SPDA, aterramento", "unid": "vb", "peso_pct": 10.0},
        {"item": "Revestimentos", "descricao": "Chapisco, reboco, cerâmica, porcelanato, pintura", "unid": "vb", "peso_pct": 12.0},
        {"item": "Esquadrias", "descricao": "Portas, janelas, vidros, ferragens", "unid": "vb", "peso_pct": 5.0},
        {"item": "Pisos", "descricao": "Contrapiso, piso cerâmico/porcelanato, rodapé", "unid": "vb", "peso_pct": 7.0},
        {"item": "Impermeabilização", "descricao": "Manta asfáltica, impermeabilização de banheiros e coberturas", "unid": "vb", "peso_pct": 3.0},
        {"item": "Serviços Complementares", "descricao": "Limpeza final, as-built, habite-se, desmobilização", "unid": "vb", "peso_pct": 2.0},
        {"item": "Ar Condicionado/HVAC", "descricao": "Sistema de climatização, dutos, splits/centrais", "unid": "vb", "peso_pct": 6.0},
    ],
    "reforma": [
        {"item": "Serviços Preliminares", "descricao": "Proteções, tapumes, sinalização, instalação canteiro", "unid": "vb", "peso_pct": 4.0},
        {"item": "Demolições e Remoções", "descricao": "Demolição de alvenaria, remoção de pisos/revestimentos, bota-fora", "unid": "vb", "peso_pct": 10.0},
        {"item": "Alvenaria e Vedação", "descricao": "Reconstrução de paredes, vergas, fechamentos", "unid": "vb", "peso_pct": 10.0},
        {"item": "Instalações Hidráulicas", "descricao": "Substituição de tubulações, louças, metais", "unid": "vb", "peso_pct": 10.0},
        {"item": "Instalações Elétricas", "descricao": "Troca de fiação, quadros, tomadas, luminárias", "unid": "vb", "peso_pct": 12.0},
        {"item": "Revestimentos e Pintura", "descricao": "Chapisco, reboco, massa corrida, pintura látex/acrílica", "unid": "vb", "peso_pct": 18.0},
        {"item": "Pisos", "descricao": "Remoção e assentamento de piso, contrapiso, rodapé", "unid": "vb", "peso_pct": 12.0},
        {"item": "Esquadrias", "descricao": "Substituição de portas, janelas, vidros, ferragens", "unid": "vb", "peso_pct": 6.0},
        {"item": "Cobertura", "descricao": "Reparo/troca de telhas, calhas, rufos, impermeabilização", "unid": "vb", "peso_pct": 8.0},
        {"item": "Impermeabilização", "descricao": "Tratamento de trincas, impermeabilização de áreas molhadas", "unid": "vb", "peso_pct": 4.0},
        {"item": "Fachada", "descricao": "Tratamento de fachada, pintura externa, pastilhas", "unid": "vb", "peso_pct": 4.0},
        {"item": "Limpeza e Entrega", "descricao": "Limpeza final, remoção de entulho, desmobilização", "unid": "vb", "peso_pct": 2.0},
    ],
    "pavimentacao": [
        {"item": "Serviços Preliminares", "descricao": "Mobilização, canteiro, sinalização, controle tecnológico", "unid": "vb", "peso_pct": 3.0},
        {"item": "Terraplenagem", "descricao": "Corte, aterro, compactação de subleito, regularização", "unid": "m³", "peso_pct": 15.0},
        {"item": "Drenagem", "descricao": "Meio-fio, sarjeta, boca-de-lobo, tubulação de drenagem", "unid": "vb", "peso_pct": 15.0},
        {"item": "Base e Sub-base", "descricao": "Sub-base granular, base de brita graduada/BGS", "unid": "m²", "peso_pct": 20.0},
        {"item": "Revestimento Asfáltico", "descricao": "Imprimação, CBUQ/TSD, compactação", "unid": "m²", "peso_pct": 25.0},
        {"item": "Sinalização", "descricao": "Sinalização horizontal e vertical, placas, faixas", "unid": "vb", "peso_pct": 5.0},
        {"item": "Calçadas e Acessibilidade", "descricao": "Calçadas em concreto, rampas, piso tátil", "unid": "vb", "peso_pct": 10.0},
        {"item": "Serviços Complementares", "descricao": "Limpeza, as-built, desmobilização", "unid": "vb", "peso_pct": 2.0},
        {"item": "Paisagismo", "descricao": "Plantio de grama/árvores, canteiros", "unid": "vb", "peso_pct": 5.0},
    ],
    "infraestrutura": [
        {"item": "Serviços Preliminares", "descricao": "Mobilização, canteiro, topografia, sondagem", "unid": "vb", "peso_pct": 3.0},
        {"item": "Terraplenagem", "descricao": "Escavação, aterro, compactação, transporte de material", "unid": "m³", "peso_pct": 15.0},
        {"item": "Drenagem Pluvial", "descricao": "Tubulação, galerias, caixas de passagem, dissipadores", "unid": "vb", "peso_pct": 20.0},
        {"item": "Rede de Esgoto", "descricao": "Tubulação PVC, PVs, ligações domiciliares", "unid": "vb", "peso_pct": 15.0},
        {"item": "Rede de Água", "descricao": "Tubulação PEAD/PVC, registros, conexões, poço", "unid": "vb", "peso_pct": 12.0},
        {"item": "Pavimentação", "descricao": "Base, sub-base, revestimento asfáltico/intertravado", "unid": "vb", "peso_pct": 18.0},
        {"item": "Iluminação Pública", "descricao": "Postes, luminárias LED, cabeamento subterrâneo", "unid": "vb", "peso_pct": 8.0},
        {"item": "Contenções", "descricao": "Muros de arrimo, gabiões, cortinas", "unid": "vb", "peso_pct": 5.0},
        {"item": "Serviços Complementares", "descricao": "Limpeza, as-built, desmobilização, paisagismo", "unid": "vb", "peso_pct": 4.0},
    ],
}

# BDI referência TCU (Acórdão 2.622/2013)
BDI_OBRA = {
    "1_minimo": {"admin_central": 3.0, "seguro": 0.8, "garantia": 0.8, "risco": 1.0, "desp_financ": 1.0, "lucro": 5.0, "tributos_pis": 0.65, "tributos_cofins": 3.0, "tributos_iss": 2.0, "tributos_cprb": 0.0},
    "2_1quartil": {"admin_central": 4.0, "seguro": 0.8, "garantia": 0.8, "risco": 1.0, "desp_financ": 1.0, "lucro": 6.5, "tributos_pis": 0.65, "tributos_cofins": 3.0, "tributos_iss": 2.0, "tributos_cprb": 0.0},
    "3_mediana": {"admin_central": 4.5, "seguro": 0.8, "garantia": 0.8, "risco": 1.2, "desp_financ": 1.2, "lucro": 7.5, "tributos_pis": 0.65, "tributos_cofins": 3.0, "tributos_iss": 3.0, "tributos_cprb": 0.0},
    "4_maximo": {"admin_central": 5.5, "seguro": 1.0, "garantia": 1.0, "risco": 1.5, "desp_financ": 1.5, "lucro": 8.5, "tributos_pis": 0.65, "tributos_cofins": 3.0, "tributos_iss": 5.0, "tributos_cprb": 0.0},
}


def _detectar_tipo_obra(objeto: str) -> str:
    """Detecta o tipo de obra pelo objeto do edital."""
    obj = objeto.lower()
    if any(k in obj for k in ["paviment", "asfalto", "recapeamento", "estrada", "rodovia"]):
        return "pavimentacao"
    if any(k in obj for k in ["drenagem", "esgoto", "saneamento", "urbanização", "urbanizacao", "infraestrutura"]):
        return "infraestrutura"
    if any(k in obj for k in ["reforma", "restaur", "recuperação", "recuperacao", "requalificação", "requalificacao", "fachada", "impermeabiliza"]):
        return "reforma"
    return "construcao"


def _calcular_bdi(cenario: dict) -> float:
    """Calcula BDI pela fórmula TCU: BDI = ((1+AC+S+R+G+DF)*(1+L)/(1-T)) - 1"""
    ac = cenario["admin_central"] / 100
    s = cenario["seguro"] / 100
    g = cenario["garantia"] / 100
    r = cenario["risco"] / 100
    df = cenario["desp_financ"] / 100
    lucro = cenario["lucro"] / 100
    t = (cenario["tributos_pis"] + cenario["tributos_cofins"] + cenario["tributos_iss"] + cenario.get("tributos_cprb", 0)) / 100
    return ((1 + ac + s + r + g + df) * (1 + lucro) / (1 - t)) - 1


def gerar_planilha_obra(
    itens: list[dict] | None = None,
    licitacao_info: dict = None,
    output_path: str | Path = "planilha_obra.xlsx",
) -> Path:
    """Gera planilha orçamentária de obra com BDI TCU.

    Args:
        itens: Lista de itens de serviço. Se None, usa template baseado no objeto.
        licitacao_info: Dict com orgao, pregao, objeto, valor_teto, prazo_meses.
        output_path: Caminho do arquivo .xlsx.
    """
    output_path = Path(output_path)
    licit = licitacao_info or {}
    objeto = licit.get("objeto", "")
    valor_teto = licit.get("valor_teto", 0)
    prazo_meses = licit.get("prazo_meses", 12)

    tipo_obra = _detectar_tipo_obra(objeto)
    if not itens:
        itens = ITENS_OBRA.get(tipo_obra, ITENS_OBRA["construcao"])

    wb = Workbook()

    # ===== ABA ORÇAMENTO =====
    ws = wb.active
    ws.title = "ORÇAMENTO"
    ws.sheet_properties.tabColor = "1F3864"

    # Larguras
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 16

    # Header
    r = 1
    for col in range(1, 9):
        _s(ws, r, col, "", fill=HEADER_FILL)
    _s(ws, r, 1, "PLANILHA ORÇAMENTÁRIA - OBRA/REFORMA", TITLE_FONT, HEADER_FILL)
    ws.merge_cells("A1:H1")

    r = 2
    _s(ws, r, 1, "Órgão:", BOLD)
    _s(ws, r, 2, licit.get("orgao", ""), NORMAL)
    ws.merge_cells("B2:D2")
    _s(ws, r, 5, "Edital:", BOLD)
    _s(ws, r, 6, licit.get("pregao", ""), NORMAL)
    ws.merge_cells("F2:H2")

    r = 3
    _s(ws, r, 1, "Objeto:", BOLD)
    _s(ws, r, 2, objeto, NORMAL, align=ALIGN_L)
    ws.merge_cells("B3:H3")

    r = 4
    _s(ws, r, 1, "Tipo:", BOLD)
    _s(ws, r, 2, tipo_obra.upper(), BOLD_BLUE)
    _s(ws, r, 3, "Referência: SINAPI", SMALL)
    _s(ws, r, 5, "Prazo:", BOLD)
    _s(ws, r, 6, f"{prazo_meses} meses", NORMAL)
    _s(ws, r, 7, "Valor Ref:", BOLD)
    _s(ws, r, 8, valor_teto, BOLD_BLUE, fmt=FMT_BRL)

    # Tabela de itens
    r = 6
    headers = ["Item", "Descrição", "Detalhamento", "Unid", "Qtd", "Preço Unit.", "Total s/ BDI", "Total c/ BDI"]
    for i, h in enumerate(headers, 1):
        _s(ws, r, i, h, BOLD_W, MOD_FILL, align=ALIGN_C)

    # BDI mediana para cálculo
    bdi_ref = _calcular_bdi(BDI_OBRA["3_mediana"])

    for idx, item in enumerate(itens):
        r += 1
        row = r
        peso = item.get("peso_pct", 100 / len(itens))

        _s(ws, row, 1, idx + 1, NORMAL, align=ALIGN_C)
        _s(ws, row, 2, item["item"], BOLD)
        _s(ws, row, 3, item.get("descricao", ""), NORMAL, align=ALIGN_L)
        _s(ws, row, 4, item.get("unid", "vb"), NORMAL, align=ALIGN_C)
        # Quantidade = 1 (verba)
        _s(ws, row, 5, 1, NORMAL, INPUT_FILL, fmt=FMT_INT, align=ALIGN_C)
        # Preço unitário = peso% * valor_teto / (1 + BDI)
        custo_direto = (peso / 100) * valor_teto / (1 + bdi_ref) if valor_teto > 0 else 0
        _s(ws, row, 6, round(custo_direto, 2), NORMAL, INPUT_FILL, fmt=FMT_BRL, align=ALIGN_R)
        # Total sem BDI = Qtd * PU
        _formula(ws, row, 7, f"=E{row}*F{row}", NORMAL, fmt=FMT_BRL)
        # Total com BDI = Total sem BDI * (1 + BDI)
        _formula(ws, row, 8, f"=G{row}*(1+BDI!B12)", BOLD, fmt=FMT_BRL)

    # Linha TOTAL
    r += 1
    total_row = r
    _s(ws, r, 1, "", fill=TOTAL_FILL)
    _s(ws, r, 2, "TOTAL GERAL", BOLD, TOTAL_FILL)
    for col in [3, 4, 5]:
        _s(ws, r, col, "", fill=TOTAL_FILL)
    _s(ws, r, 6, "", fill=TOTAL_FILL)
    _formula(ws, r, 7, f"=SUM(G7:G{r-1})", BOLD, TOTAL_FILL, FMT_BRL)
    _formula(ws, r, 8, f"=SUM(H7:H{r-1})", BOLD, TOTAL_FILL, FMT_BRL)

    # Linha comparação com teto
    r += 1
    _s(ws, r, 2, "VALOR REFERÊNCIA (TETO)", BOLD)
    _s(ws, r, 8, valor_teto, BOLD, fmt=FMT_BRL)

    r += 1
    _s(ws, r, 2, "DESCONTO SOBRE TETO", BOLD_BLUE)
    if valor_teto > 0:
        _formula(ws, r, 8, f"=1-(H{total_row}/H{r-1})", BOLD_BLUE, fmt=FMT_PCT)
    else:
        _s(ws, r, 8, 0, BOLD_BLUE, fmt=FMT_PCT)

    r += 2
    _s(ws, r, 2, "IMPORTANTE:", RED_BOLD)
    _s(ws, r, 3, "Valores estimados. Substituir pelos preços SINAPI/SICRO reais da composição de custos.", RED_BOLD)
    ws.merge_cells(f"C{r}:H{r}")

    # ===== ABA BDI =====
    ws_bdi = wb.create_sheet("BDI")
    ws_bdi.sheet_properties.tabColor = "4472C4"

    ws_bdi.column_dimensions["A"].width = 5
    ws_bdi.column_dimensions["B"].width = 35
    ws_bdi.column_dimensions["C"].width = 14
    ws_bdi.column_dimensions["D"].width = 14
    ws_bdi.column_dimensions["E"].width = 14
    ws_bdi.column_dimensions["F"].width = 14
    ws_bdi.column_dimensions["G"].width = 20

    r = 1
    for col in range(1, 8):
        _s(ws_bdi, r, col, "", fill=HEADER_FILL)
    _s(ws_bdi, r, 1, "COMPOSIÇÃO DO BDI — REFERÊNCIA TCU (Acórdão 2.622/2013)", TITLE_FONT, HEADER_FILL)
    ws_bdi.merge_cells("A1:G1")

    r = 3
    cols = ["", "Componente", "Mínimo", "1º Quartil", "Mediana (Usado)", "Máximo", "Seu BDI"]
    for i, h in enumerate(cols, 1):
        _s(ws_bdi, r, i, h, BOLD_W, MOD_FILL, align=ALIGN_C)

    componentes = [
        ("Administração Central", "admin_central"),
        ("Seguro + Garantia", None),
        ("  Seguro", "seguro"),
        ("  Garantia", "garantia"),
        ("Risco", "risco"),
        ("Despesas Financeiras", "desp_financ"),
        ("Lucro", "lucro"),
        ("Tributos", None),
        ("  PIS", "tributos_pis"),
        ("  COFINS", "tributos_cofins"),
        ("  ISS", "tributos_iss"),
    ]

    start_row = 4
    for idx, (nome, key) in enumerate(componentes):
        r = start_row + idx
        _s(ws_bdi, r, 1, idx + 1 if key else "", NORMAL, align=ALIGN_C)
        _s(ws_bdi, r, 2, nome, BOLD if key is None else NORMAL)
        if key:
            for ci, cenario_key in enumerate(["1_minimo", "2_1quartil", "3_mediana", "4_maximo"]):
                val = BDI_OBRA[cenario_key][key] / 100
                _s(ws_bdi, r, 3 + ci, val, NORMAL, fmt=FMT_PCT, align=ALIGN_C)
            # Coluna "Seu BDI" — editável, começa com mediana
            _s(ws_bdi, r, 7, BDI_OBRA["3_mediana"][key] / 100, NORMAL, INPUT_FILL, FMT_PCT, ALIGN_C)

    # Linha BDI calculado
    r = start_row + len(componentes) + 1
    bdi_row = r
    _s(ws_bdi, r, 1, "", fill=TOTAL_FILL)
    _s(ws_bdi, r, 2, "BDI CALCULADO", BOLD, TOTAL_FILL)

    # Mapear rows das componentes para a fórmula
    # AC=row4, Seg=row6, Gar=row7, Risco=row8, DF=row9, Lucro=row10, PIS=row12, COFINS=row13, ISS=row14
    ac_r, seg_r, gar_r, risco_r, df_r, lucro_r, pis_r, cofins_r, iss_r = 4, 6, 7, 8, 9, 10, 12, 13, 14

    for ci, cenario_key in enumerate(["1_minimo", "2_1quartil", "3_mediana", "4_maximo"]):
        col = 3 + ci
        bdi_val = _calcular_bdi(BDI_OBRA[cenario_key])
        _s(ws_bdi, r, col, bdi_val, BOLD, TOTAL_FILL, FMT_PCT, ALIGN_C)

    # Fórmula BDI na coluna "Seu BDI"
    # BDI = ((1+AC+S+G+R+DF)*(1+L)/(1-(PIS+COFINS+ISS))) - 1
    formula_bdi = (
        f"=((1+G{ac_r}+G{seg_r}+G{gar_r}+G{risco_r}+G{df_r})"
        f"*(1+G{lucro_r})"
        f"/(1-(G{pis_r}+G{cofins_r}+G{iss_r})))-1"
    )
    _formula(ws_bdi, r, 7, formula_bdi, BOLD, SUBTOTAL_FILL, FMT_PCT)

    # Named range para BDI usado na aba ORÇAMENTO
    # Referência: BDI!B12 -> vamos colocar o BDI mediana numa célula de referência
    r += 2
    _s(ws_bdi, r, 1, "", fill=SUBTOTAL_FILL)
    _s(ws_bdi, r, 2, "BDI ADOTADO (usado no orçamento) →", BOLD, SUBTOTAL_FILL)
    _formula(ws_bdi, r, 3, f"=G{bdi_row}", BOLD_BLUE, SUBTOTAL_FILL, FMT_PCT)
    # Guardar a referência - aba ORÇAMENTO usa BDI!B12 na fórmula
    # Mas precisamos que o BDI fique numa célula conhecida
    # Vou colocar em B{bdi_row} a referência que ORÇAMENTO usa
    bdi_ref_row = r

    # Corrigir referência na aba ORÇAMENTO: em vez de BDI!B12, usa BDI!C{bdi_ref_row}
    # Preciso voltar e ajustar as fórmulas da aba ORÇAMENTO
    ws_orc = wb["ORÇAMENTO"]
    for row_idx in range(7, total_row):
        cell = ws_orc.cell(row=row_idx, column=8)
        if cell.value and isinstance(cell.value, str) and "BDI!B12" in cell.value:
            cell.value = cell.value.replace("BDI!B12", f"BDI!C{bdi_ref_row}")

    # ===== ABA CENÁRIOS =====
    ws_cen = wb.create_sheet("CENÁRIOS")
    ws_cen.sheet_properties.tabColor = "E2EFDA"

    ws_cen.column_dimensions["A"].width = 5
    ws_cen.column_dimensions["B"].width = 22
    ws_cen.column_dimensions["C"].width = 14
    ws_cen.column_dimensions["D"].width = 18
    ws_cen.column_dimensions["E"].width = 18
    ws_cen.column_dimensions["F"].width = 14
    ws_cen.column_dimensions["G"].width = 14

    r = 1
    for col in range(1, 8):
        _s(ws_cen, r, col, "", fill=HEADER_FILL)
    _s(ws_cen, r, 1, "SIMULAÇÃO DE CENÁRIOS DE LANCE", TITLE_FONT, HEADER_FILL)
    ws_cen.merge_cells("A1:G1")

    r = 3
    for i, h in enumerate(["#", "Cenário", "BDI", "Valor Proposta", "Valor Teto", "Desconto", "Situação"], 1):
        _s(ws_cen, r, i, h, BOLD_W, MOD_FILL, align=ALIGN_C)

    cenarios = [
        ("Conservador (Máximo)", "4_maximo"),
        ("Moderado (Mediana)", "3_mediana"),
        ("Competitivo (1º Quartil)", "2_1quartil"),
        ("Agressivo (Mínimo)", "1_minimo"),
    ]

    custo_direto_total = sum(
        (item.get("peso_pct", 100 / len(itens)) / 100) * valor_teto / (1 + bdi_ref) if valor_teto > 0 else 0
        for item in itens
    )

    for idx, (nome, key) in enumerate(cenarios):
        r = 4 + idx
        bdi_val = _calcular_bdi(BDI_OBRA[key])
        proposta = custo_direto_total * (1 + bdi_val)

        _s(ws_cen, r, 1, idx + 1, NORMAL, align=ALIGN_C)
        _s(ws_cen, r, 2, nome, BOLD)
        _s(ws_cen, r, 3, bdi_val, NORMAL, fmt=FMT_PCT, align=ALIGN_C)
        _s(ws_cen, r, 4, proposta, BOLD, fmt=FMT_BRL)
        _s(ws_cen, r, 5, valor_teto, NORMAL, fmt=FMT_BRL)
        desc = (1 - proposta / valor_teto) if valor_teto > 0 else 0
        _s(ws_cen, r, 6, desc, NORMAL, fmt=FMT_PCT, align=ALIGN_C)
        situacao = "OK" if proposta <= valor_teto else "ACIMA DO TETO"
        fill = SUBTOTAL_FILL if proposta <= valor_teto else ALERT_FILL
        _s(ws_cen, r, 7, situacao, BOLD, fill, align=ALIGN_C)

    # ===== ABA CRONOGRAMA =====
    ws_cron = wb.create_sheet("CRONOGRAMA")
    ws_cron.sheet_properties.tabColor = "FFB038"

    ws_cron.column_dimensions["A"].width = 6
    ws_cron.column_dimensions["B"].width = 28

    r = 1
    max_col = min(prazo_meses, 24) + 3
    for col in range(1, max_col + 1):
        _s(ws_cron, r, col, "", fill=HEADER_FILL)
    _s(ws_cron, r, 1, "CRONOGRAMA FÍSICO-FINANCEIRO", TITLE_FONT, HEADER_FILL)
    ws_cron.merge_cells(f"A1:{chr(64+max_col)}1")

    r = 3
    _s(ws_cron, r, 1, "Item", BOLD_W, MOD_FILL, align=ALIGN_C)
    _s(ws_cron, r, 2, "Serviço", BOLD_W, MOD_FILL, align=ALIGN_C)
    meses_show = min(prazo_meses, 24)
    for m in range(1, meses_show + 1):
        col = m + 2
        ws_cron.column_dimensions[chr(64 + col) if col <= 26 else ""].width = 12
        _s(ws_cron, r, col, f"Mês {m}", BOLD_W, MOD_FILL, align=ALIGN_C)

    _s(ws_cron, r, meses_show + 3, "Total %", BOLD_W, MOD_FILL, align=ALIGN_C)

    for idx, item in enumerate(itens):
        r = 4 + idx
        _s(ws_cron, r, 1, idx + 1, NORMAL, align=ALIGN_C)
        _s(ws_cron, r, 2, item["item"], NORMAL)
        peso = item.get("peso_pct", 0)

        # Distribui o peso uniformemente nos meses (placeholder editável)
        pct_mes = peso / meses_show / 100 if meses_show > 0 else 0
        for m in range(1, meses_show + 1):
            _s(ws_cron, r, m + 2, pct_mes, NORMAL, INPUT_FILL, FMT_PCT, ALIGN_C)

        # Total = soma dos meses
        col_ini = chr(67)  # C
        # Para mais de 24 meses precisaria de outra abordagem
        last_col_letter = chr(66 + meses_show)
        _formula(ws_cron, r, meses_show + 3, f"=SUM(C{r}:{last_col_letter}{r})", BOLD, fmt=FMT_PCT)

    # Linha total por mês
    r = 4 + len(itens)
    _s(ws_cron, r, 2, "TOTAL MÊS (%)", BOLD, TOTAL_FILL)
    for m in range(1, meses_show + 1):
        col = m + 2
        col_letter = chr(64 + col)
        _formula(ws_cron, r, col, f"=SUM({col_letter}4:{col_letter}{r-1})", BOLD, TOTAL_FILL, FMT_PCT)
    _formula(ws_cron, r, meses_show + 3, f"=SUM(C{r}:{chr(66+meses_show)}{r})", BOLD, TOTAL_FILL, FMT_PCT)

    # ===== ABA ENCARGOS SOCIAIS =====
    ws_enc = wb.create_sheet("ENCARGOS SOCIAIS")
    ws_enc.sheet_properties.tabColor = "A78BFA"

    ws_enc.column_dimensions["A"].width = 5
    ws_enc.column_dimensions["B"].width = 40
    ws_enc.column_dimensions["C"].width = 16

    r = 1
    for col in range(1, 4):
        _s(ws_enc, r, col, "", fill=HEADER_FILL)
    _s(ws_enc, r, 1, "ENCARGOS SOCIAIS SOBRE MÃO DE OBRA (OBRA)", TITLE_FONT, HEADER_FILL)
    ws_enc.merge_cells("A1:C1")

    encargos = [
        ("", "GRUPO A — Encargos Sociais Básicos", None),
        ("A1", "INSS", 20.0),
        ("A2", "SESI/SESC", 1.5),
        ("A3", "SENAI/SENAC", 1.0),
        ("A4", "INCRA", 0.2),
        ("A5", "Salário Educação", 2.5),
        ("A6", "FGTS", 8.0),
        ("A7", "Seguro Acidente de Trabalho (SAT/RAT)", 3.0),
        ("A8", "SEBRAE", 0.6),
        ("", "TOTAL GRUPO A", None),
        ("", "", None),
        ("", "GRUPO B — Encargos que recebem incidência de A", None),
        ("B1", "13º Salário", 8.33),
        ("B2", "Férias (1/3 constitucional)", 2.78),
        ("B3", "Auxílio Doença (15 dias)", 0.55),
        ("B4", "Faltas Justificadas", 0.68),
        ("B5", "Licença Paternidade", 0.02),
        ("B6", "Acidente de Trabalho (15 dias)", 0.11),
        ("", "TOTAL GRUPO B", None),
        ("", "", None),
        ("", "GRUPO C — Encargos que não recebem incidência", None),
        ("C1", "Aviso Prévio Indenizado", 4.11),
        ("C2", "Multa FGTS (40%)", 3.20),
        ("C3", "Indenização Adicional", 0.42),
        ("", "TOTAL GRUPO C", None),
        ("", "", None),
        ("", "GRUPO D — Reincidências", None),
        ("D1", "Incidência de A sobre B", None),
        ("", "TOTAL GRUPO D", None),
        ("", "", None),
        ("", "TOTAL ENCARGOS SOCIAIS (A+B+C+D)", None),
    ]

    r = 3
    _s(ws_enc, r, 1, "Item", BOLD_W, MOD_FILL, align=ALIGN_C)
    _s(ws_enc, r, 2, "Descrição", BOLD_W, MOD_FILL)
    _s(ws_enc, r, 3, "%", BOLD_W, MOD_FILL, align=ALIGN_C)

    grupo_a_start = 0
    grupo_a_end = 0
    grupo_b_start = 0
    grupo_b_end = 0
    grupo_c_start = 0
    grupo_c_end = 0

    for idx, (cod, desc, val) in enumerate(encargos):
        r = 4 + idx
        _s(ws_enc, r, 1, cod, NORMAL, align=ALIGN_C)

        if desc.startswith("GRUPO") or desc.startswith("TOTAL"):
            _s(ws_enc, r, 2, desc, BOLD, TOTAL_FILL if "TOTAL" in desc else MOD_FILL,
               align=ALIGN_L)
            if "TOTAL" in desc:
                _s(ws_enc, r, 3, "", BOLD, TOTAL_FILL, FMT_PCT, ALIGN_C)
        elif val is not None:
            _s(ws_enc, r, 2, desc, NORMAL)
            _s(ws_enc, r, 3, val / 100, NORMAL, INPUT_FILL, FMT_PCT, ALIGN_C)
        else:
            _s(ws_enc, r, 2, desc, NORMAL)

    # Salvar
    wb.save(str(output_path))
    log.info(f"Planilha de obra gerada: {output_path}")
    return output_path
