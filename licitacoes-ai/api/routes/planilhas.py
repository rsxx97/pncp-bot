"""API de análise de planilhas — lê xlsx/ods locais sem custo de API."""
import json
import logging
from pathlib import Path

from fastapi import APIRouter
import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger("api.planilhas")

router = APIRouter(prefix="/api/planilhas", tags=["planilhas"])

DATA_DIR = Path(__file__).parent.parent.parent / "data"
PLANILHAS_DIR = DATA_DIR / "planilhas"
EDITAIS_DIR = DATA_DIR / "editais"


def _ler_planilha(path: Path) -> dict:
    """Lê um xlsx e extrai resumo, cargos e cenários."""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        return {"arquivo": path.name, "erro": str(e)}

    resultado = {
        "arquivo": path.name,
        "tamanho_kb": round(path.stat().st_size / 1024, 1),
        "abas": wb.sheetnames,
        "cargos": [],
        "resumo": None,
        "break_even": None,
        "cenarios": [],
    }

    # Extrair dados de cada aba
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nome_upper = sheet_name.upper()

        if nome_upper == "RESUMO":
            resultado["resumo"] = _extrair_resumo(ws)
        elif nome_upper == "BREAK-EVEN":
            be = _extrair_breakeven(ws)
            resultado["break_even"] = be.get("break_even")
            resultado["cenarios"] = be.get("cenarios", [])
        elif nome_upper not in ("BREAK-EVEN",):
            cargo = _extrair_cargo(ws, sheet_name)
            if cargo and cargo.get("salario_base"):
                resultado["cargos"].append(cargo)

    wb.close()
    return resultado


def _extrair_cargo(ws, nome_aba: str) -> dict | None:
    """Extrai dados de uma aba de cargo (Módulo IN 05/2017)."""
    cargo = {"nome": nome_aba, "salario_base": 0, "postos": 0, "jornada": "", "cbo": "", "cct": ""}

    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v).strip()

            # Salário
            col_letter = get_column_letter(cell.column)
            row_num = cell.row
            if col_letter == "B" and row_num == 5:
                try:
                    cargo["salario_base"] = float(v)
                except (ValueError, TypeError):
                    pass
            # Postos
            if col_letter == "B" and row_num == 4:
                try:
                    cargo["postos"] = int(v)
                except (ValueError, TypeError):
                    pass
            # Jornada
            if col_letter == "D" and row_num == 4:
                cargo["jornada"] = sv
            # CBO
            if col_letter == "D" and row_num == 3:
                cargo["cbo"] = sv
            # CCT
            if col_letter == "B" and row_num == 2:
                cargo["cct"] = sv
            # Cargo nome (coluna B, row 3)
            if col_letter == "B" and row_num == 3:
                cargo["nome"] = sv

    return cargo if cargo["salario_base"] > 0 else None


def _extrair_resumo(ws) -> dict:
    """Extrai aba RESUMO."""
    resumo = {
        "titulo": "",
        "regime": "",
        "itens": [],
        "total_postos": 0,
        "total_mensal": 0,
        "total_anual": 0,
        "teto_edital": 0,
        "proposta": 0,
        "desconto_pct": 0,
        "margem": 0,
    }

    # Row 1 = título, Row 2 = regime/CCT
    r1 = ws.cell(1, 1).value
    r2 = ws.cell(2, 1).value
    if r1:
        resumo["titulo"] = str(r1)
    if r2:
        resumo["regime"] = str(r2)

    # Itens de cargo (a partir da row 5 normalmente)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        vals = {get_column_letter(c.column): c.value for c in row if c.value is not None}

        # Linha de item (tem número em A e cargo em B)
        if "A" in vals and "B" in vals:
            try:
                item_num = int(vals["A"])
                item = {
                    "cargo": str(vals["B"]),
                    "postos": int(vals.get("C", 0)) if vals.get("C") else 0,
                    "valor_emp_mes": float(vals.get("D", 0)) if vals.get("D") else 0,
                    "valor_mensal": float(vals.get("E", 0)) if vals.get("E") else 0,
                    "valor_anual": float(vals.get("F", 0)) if vals.get("F") else 0,
                }
                resumo["itens"].append(item)
            except (ValueError, TypeError):
                pass

        # Linha TOTAL
        if "B" in vals and "TOTAL" in str(vals["B"]).upper():
            resumo["total_postos"] = int(vals.get("C", 0)) if vals.get("C") else 0
            resumo["total_mensal"] = float(vals.get("E", 0)) if vals.get("E") else 0
            resumo["total_anual"] = float(vals.get("F", 0)) if vals.get("F") else 0

        # Teto edital
        if "B" in vals and "TETO" in str(vals["B"]).upper():
            resumo["teto_edital"] = float(vals.get("F", 0)) if vals.get("F") else 0

        # Proposta
        if "B" in vals and "PROPOSTA" in str(vals["B"]).upper():
            resumo["proposta"] = float(vals.get("F", 0)) if vals.get("F") else 0

        # Desconto
        if "B" in vals and "DESCONTO" in str(vals["B"]).upper():
            v = vals.get("F", 0)
            if v:
                try:
                    fv = float(v) if not isinstance(v, str) else float(v.replace("%", "")) / 100
                    resumo["desconto_pct"] = fv
                except (ValueError, TypeError):
                    pass

        # Margem
        if "B" in vals and "MARGEM" in str(vals["B"]).upper():
            resumo["margem"] = float(vals.get("F", 0)) if vals.get("F") else 0

    return resumo


def _extrair_breakeven(ws) -> dict:
    """Extrai aba BREAK-EVEN."""
    result = {"break_even": 0, "cenarios": []}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = {get_column_letter(c.column): c.value for c in row if c.value is not None}

        # Break-even total
        if "B" in vals and "BREAK" in str(vals["B"]).upper() and "CUSTO" in str(vals["B"]).upper():
            result["break_even"] = float(vals.get("I", vals.get("F", 0)) or 0)

        # Cenários
        if "A" in vals and "B" in vals:
            try:
                int(vals["A"])  # É número = cenário
                nome_cenario = str(vals["B"])
                if any(k in nome_cenario.lower() for k in ["conserv", "moder", "agress", "mín", "min"]):
                    cenario = {
                        "nome": nome_cenario,
                        "ci_pct": str(vals.get("C", "")),
                        "lucro_pct": str(vals.get("D", "")),
                        "valor_anual": float(vals.get("E", 0)) if vals.get("E") else 0,
                        "desconto_teto": str(vals.get("F", "")),
                        "margem_bruta": float(vals.get("G", 0)) if vals.get("G") else 0,
                        "status": str(vals.get("I", "")),
                    }
                    result["cenarios"].append(cenario)
            except (ValueError, TypeError):
                pass

    return result


@router.get("/resumo")
def listar_planilhas():
    """Lista todas as planilhas com análise detalhada."""
    planilhas = []

    # Planilhas geradas
    if PLANILHAS_DIR.exists():
        for f in sorted(PLANILHAS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
            if f.name.startswith("teste_"):
                continue  # Pula arquivos de teste
            planilhas.append(_ler_planilha(f))

    # Planilhas na raiz de data/
    if DATA_DIR.exists():
        for f in sorted(DATA_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
            if f.name.startswith("teste_"):
                continue
            planilhas.append(_ler_planilha(f))

    # Planilhas do edital (templates do órgão)
    if EDITAIS_DIR.exists():
        for f in sorted(EDITAIS_DIR.glob("*lanilha*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
            info = _ler_planilha(f)
            info["tipo"] = "template_orgao"
            planilhas.append(info)

    return {"planilhas": planilhas, "total": len(planilhas)}


@router.get("/resumo/{arquivo}")
def detalhe_planilha(arquivo: str):
    """Retorna análise detalhada de uma planilha específica."""
    # Busca em planilhas e editais
    for d in [PLANILHAS_DIR, EDITAIS_DIR]:
        path = d / arquivo
        if path.exists():
            return _ler_planilha(path)
    return {"erro": "Arquivo não encontrado"}


@router.get("/verificacao/ultima")
def get_verificacao():
    """Retorna última verificação cruzada das planilhas."""
    verif_file = DATA_DIR.parent / "bot_pncp_skills" / "data" / "verificacoes" / "ultima_verificacao.json"
    if not verif_file.exists():
        return {"erro": "Nenhuma verificação disponível. Execute o bot verificador."}
    return json.loads(verif_file.read_text(encoding="utf-8"))


@router.post("/verificacao/executar")
def executar_verificacao():
    """Dispara bot verificador em background."""
    import threading, sys
    def _run():
        sys.path.insert(0, str(DATA_DIR.parent))
        from bot_pncp_skills.bot_verificador import verificar_tudo
        verificar_tudo()
    threading.Thread(target=_run, daemon=True).start()
    return {"ok": True, "msg": "Verificação iniciada em background"}


@router.get("/skills/atual")
def get_skills():
    """Retorna skills consolidadas das planilhas absorvidas do PNCP."""
    skills_file = DATA_DIR.parent / "bot_pncp_skills" / "data" / "skills" / "skills_consolidado.json"
    if not skills_file.exists():
        return {"erro": "Skills não disponíveis."}
    return json.loads(skills_file.read_text(encoding="utf-8"))
